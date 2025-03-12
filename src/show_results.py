from IPython.display import display
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from tabulate import tabulate
from .functions import get_month_start_end

import folium
import geopandas as gpd
import glob
import os
import pandas as pd
import pickle


def plot_clusters_with_folium(
    df_scanning_locations, points_size=5, noise_size=5, polyline_points_list=[]
):

    gdf = gpd.GeoDataFrame(
        df_scanning_locations,
        geometry=gpd.points_from_xy(
            df_scanning_locations.LONGITUDE, df_scanning_locations.LATITUDE
        ),
    )  # x -> longitude ; y -> latitude

    m = folium.Map(
        location=[gdf["LATITUDE"].mean(), gdf["LONGITUDE"].mean()],
        tiles="https://wprd01.is.autonavi.com/appmaptile?x={x}&y={y}&z={z}&lang=zh_cn&size=1&scl=1&style=7",
        attr="高德-常规图",
        zoom_start=7,
    )

    unique_clusters = gdf["cluster_label"].sort_values().unique()

    colors = [
        "#1f77b4",
        "#ff7f0e",
        "#2ca02c",
        "#aab7ff",
        "#9467bd",
        "#8c564b",
        "#e377c2",
        "#7f7f7f",
        "#bcbd22",
        "#17becf",
        "#ffbb78",
        "#98df8a",
        "#ff9896",
        "#c5b0d5",
        "#c49c94",
        "#f7b6d2",
        "#dbdb8d",
        "#9edae5",
        "#f5b041",
        "#d62728",
    ]

    for i, cluster in enumerate(unique_clusters):
        if cluster != -1:  # 排除噪声点
            cluster_points = gdf[gdf["cluster_label"] == cluster]

            color_index = cluster % 20
            color = colors[color_index]
            for _, row in cluster_points.iterrows():

                folium.CircleMarker(
                    location=(row["LATITUDE"], row["LONGITUDE"]),
                    radius=points_size,
                    color=color,
                    fill=True,
                    fill_color=color,
                    fill_opacity=0.6,
                    popup=f"Cluster: {cluster}",
                ).add_to(m)

    # 绘制噪声点（可选）
    noise_points = gdf[gdf["cluster_label"] == -1]
    for _, row in noise_points.iterrows():
        folium.CircleMarker(
            location=(row["LATITUDE"], row["LONGITUDE"]),
            radius=noise_size,
            color="red",
            fill=True,
            fill_color="red",
            fill_opacity=0.6,
            popup="Noise",
        ).add_to(m)

    legend_html = """
    <div style="
        position: fixed; 
        top: 0px; left: 0px; 
        width: 120px; height: auto; 
        border: 1px solid grey; 
        z-index: 9999; font-size: 12px; 
        background-color: rgba(255, 255, 255, 0.6); 
        padding: 2px; 
        border-radius: 8px; 
        box-shadow: 0 2px 6px rgba(0,0,0,0.3); 
        overflow-y: auto; max-height: 200px;">
        <b>Cluster Legend</b><br>
    """
    for i, cluster in enumerate(unique_clusters):
        if cluster != -1:
            color_index = cluster % 20
            color = colors[color_index]
            legend_html += f'<i style="background:{color}; width: 15px; height: 15px; display:inline-block; margin-right: 5px;"></i> Cluster {cluster}<br>'
    legend_html += "</div>"

    # 将图例添加到地图
    folium.Marker(
        location=[gdf["LATITUDE"].mean(), gdf["LONGITUDE"].mean()],
        icon=folium.DivIcon(html=legend_html),
    ).add_to(m)

    if polyline_points_list:
        for polyline_points in polyline_points_list:
            folium.PolyLine(
                locations=polyline_points, color="BLUE", weight=2.5, opacity=1
            ).add_to(m)

    return m


def show_region_short_results(
    df_dealer_results,
    df_total_scanning_locations,
    start_date_str,
    end_date_str,
    dealer_region_name,
):

    df_suspicious_dealers = (
        df_dealer_results.loc[df_dealer_results["is_dealer_suspicious"] == 1, :]
        .sort_values(by="BELONG_DEALER_NO")
        .reset_index(drop=True)
    )

    product_group_name = df_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]

    # Output 1
    title_1_str = (
        "-" * 30
        + f"基于模型v1.0, ({start_date_str} - {end_date_str}),  ({dealer_region_name} - {product_group_name})的结果如下"
        + "-" * 30
    )
    # if print_title:
    #     print(title_1_str)
    print()

    # Output2
    title_2_str = "经销商数量统计"
    print(title_2_str)
    # display_fifth_title(title_2_str)
    df_region_dealer_statistics = pd.DataFrame(
        {
            "扫码经销商总数": df_total_scanning_locations.BELONG_DEALER_NO.nunique(),
            "经营范围未归档经销商数量": df_total_scanning_locations.loc[
                df_total_scanning_locations["is_dealer_within_archive"] == 0, :
            ].BELONG_DEALER_NO.nunique(),
            "当前规则下可疑经销商数量": df_suspicious_dealers.shape[0],
        },
        index=[0],
    )
    output_2 = df_region_dealer_statistics.copy()
    print(tabulate(output_2, headers="keys", tablefmt="pretty", showindex=False))
    print()

    # Output3
    title_3_str = "可疑经销商汇总表"
    print(title_3_str)
    suspicious_dealers_overall_cols = [
        "BELONG_DEALER_NO",
        "BELONG_DEALER_NAME",
        "PRODUCT_GROUP_NAME",
        "is_dealer_no_valid_scope",
        "dealer_suspicious_points_count",
        "dealer_suspicious_hotspot_count",
        "dealer_suspicious_points_ratio",
        "dealer_remote_ratio",
        "dealer_total_scanning_count",
        # "border_scanning_ratio"
    ]
    rename_dict = {
        "BELONG_DEALER_NO": "经销商编码",
        "BELONG_DEALER_NAME": "经销商名称",
        "PRODUCT_GROUP_NAME": "品项名称",
        "is_dealer_no_valid_scope": "无<有效>经营范围",
        "dealer_suspicious_points_count": "可疑扫码数量",
        "dealer_suspicious_hotspot_count": "可疑扫码热点数量",
        "dealer_suspicious_points_ratio": "扫码可疑率",
        "dealer_remote_ratio": "扫码异地率",
        "dealer_total_scanning_count": "总扫码量",
        # "border_scanning_ratio": "近边界扫码率"
    }

    df_suspicious_dealers_to_show = df_suspicious_dealers.sort_values(
        by=["is_dealer_no_valid_scope", "BELONG_DEALER_NO"], ascending=[True, True]
    ).reset_index(drop=True)

    df_suspicious_dealers_to_show["is_dealer_no_valid_scope"] = (
        df_suspicious_dealers_to_show["is_dealer_no_valid_scope"].map(
            {1: "是", 0: "否"}
        )
    )

    # 这里可以增加计算边界开瓶率？？？？？
    # df_suspicious_dealers_to_show = add_border_scanning_ratio_to_df_suspicious_dealers(
    #     df_suspicious_dealers_to_show, df_total_scanning_locations
    # )
    df_suspicious_dealers_to_show = df_suspicious_dealers_to_show.loc[
        :, suspicious_dealers_overall_cols
    ]
    df_suspicious_dealers_to_show = df_suspicious_dealers_to_show.rename(
        columns=rename_dict
    )
    output_3 = df_suspicious_dealers_to_show.copy()
    output_3[["扫码可疑率", "扫码异地率"]] = output_3[
        ["扫码可疑率", "扫码异地率"]
    ].round(3)
    print(tabulate(output_3, headers="keys", tablefmt="pretty", showindex=False))
    print()

    # Output 4
    title_4_str = "经营范围未归档经销商"
    print(title_4_str)
    unarchive_dealers_overall_cols = [
        "BELONG_DEALER_NO",
        "BELONG_DEALER_NAME",
        "PRODUCT_GROUP_NAME",
    ]
    df_unarchive_dealers_to_show = df_total_scanning_locations.loc[
        df_total_scanning_locations["is_dealer_within_archive"] == 0,
        unarchive_dealers_overall_cols,
    ].drop_duplicates()
    df_unarchive_dealers_to_show = df_unarchive_dealers_to_show.sort_values(
        by="BELONG_DEALER_NO"
    ).reset_index(drop=True)
    df_unarchive_dealers_to_show = df_unarchive_dealers_to_show.rename(
        columns=rename_dict
    )
    output_4 = df_unarchive_dealers_to_show.copy()
    print(tabulate(output_4, headers="keys", tablefmt="pretty", showindex=False))
    print()

    print("*" * 100)
    return (
        title_1_str,
        title_2_str,
        output_2,
        title_3_str,
        output_3,
        title_4_str,
        output_4,
    )


def show_region_short_results_special(
    df_dealer_results,
    df_total_scanning_locations,
    start_date_str,
    end_date_str,
    dealer_region_name,
):

    df_suspicious_dealers = (
        df_dealer_results.loc[df_dealer_results["is_dealer_suspicious_final"] == 1, :]
        .sort_values(by="BELONG_DEALER_NO")
        .reset_index(drop=True)
    )
    product_group_name = df_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]

    # Output 1
    title_1_str = (
        "-" * 30
        + f"基于模型v1.0, ({start_date_str} - {end_date_str}),  ({dealer_region_name} - {product_group_name})的结果（简略）如下"
        + "-" * 30
    )
    # if print_title:
    #     print(title_1_str)
    print()

    # Output_2
    title_2_str = "经销商数量统计"
    print(title_2_str)
    df_region_dealer_statistics = pd.DataFrame(
        {
            "扫码经销商总数": df_total_scanning_locations.BELONG_DEALER_NO.nunique(),
            "经营范围未归档经销商数量": df_total_scanning_locations.loc[
                df_total_scanning_locations["is_dealer_within_archive"] == 0, :
            ].BELONG_DEALER_NO.nunique(),
            # "当前规则下可疑经销商数量": len(ids_suspicious_total),
            "当前规则下可疑经销商数量": df_suspicious_dealers.shape[0],
        },
        index=[0],
    )
    output_2 = df_region_dealer_statistics.copy()
    print(tabulate(output_2, headers="keys", tablefmt="pretty", showindex=False))
    print()

    # Output_3
    title_3_str = "可疑经销商汇总表"
    print(title_3_str)
    suspicious_dealers_overall_cols = [
        "BELONG_DEALER_NO",
        "BELONG_DEALER_NAME",
        "PRODUCT_GROUP_NAME",
        "is_dealer_no_valid_scope",
        "dealer_suspicious_points_count_final",
        "dealer_suspicious_hotspot_count_final",
        "dealer_suspicious_points_ratio_final",
        "dealer_remote_ratio",
        "dealer_total_scanning_count",
        # "border_scanning_ratio"
    ]
    rename_dict = {
        "BELONG_DEALER_NO": "经销商编码",
        "BELONG_DEALER_NAME": "经销商名称",
        "PRODUCT_GROUP_NAME": "品项名称",
        "is_dealer_no_valid_scope": "无<有效>经营范围",
        "dealer_suspicious_points_count_final": "可疑扫码数量",
        "dealer_suspicious_hotspot_count_final": "可疑扫码热点数量",
        "dealer_suspicious_points_ratio_final": "扫码可疑率",
        "dealer_remote_ratio": "扫码异地率",
        "dealer_total_scanning_count": "总扫码量",
        # "border_scanning_ratio": "近边界扫码率"
    }
    df_suspicious_dealers_to_show = df_suspicious_dealers.sort_values(
        by=["is_dealer_no_valid_scope", "BELONG_DEALER_NO"], ascending=[True, True]
    ).reset_index(drop=True)
    df_suspicious_dealers_to_show["is_dealer_no_valid_scope"] = (
        df_suspicious_dealers_to_show["is_dealer_no_valid_scope"].map(
            {1: "是", 0: "否"}
        )
    )

    # df_suspicious_dealers_to_show = add_border_scanning_ratio_to_df_suspicious_dealers(
    #     df_suspicious_dealers_to_show, df_total_scanning_locations
    # )
    df_suspicious_dealers_to_show = df_suspicious_dealers_to_show.loc[
        :, suspicious_dealers_overall_cols
    ]
    df_suspicious_dealers_to_show = df_suspicious_dealers_to_show.rename(
        columns=rename_dict
    )
    output_3 = df_suspicious_dealers_to_show.copy()
    output_3[["扫码可疑率", "扫码异地率"]] = output_3[
        ["扫码可疑率", "扫码异地率"]
    ].round(3)
    print(tabulate(output_3, headers="keys", tablefmt="pretty", showindex=False))
    print()

    # Output 4
    title_4_str = "经营范围未归档经销商"
    print(title_4_str)
    unarchive_dealers_overall_cols = [
        "BELONG_DEALER_NO",
        "BELONG_DEALER_NAME",
        "PRODUCT_GROUP_NAME",
    ]
    df_unarchive_dealers_to_show = df_total_scanning_locations.loc[
        df_total_scanning_locations["is_dealer_within_archive"] == 0,
        unarchive_dealers_overall_cols,
    ].drop_duplicates()
    df_unarchive_dealers_to_show = df_unarchive_dealers_to_show.sort_values(
        by="BELONG_DEALER_NO"
    ).reset_index(drop=True)
    df_unarchive_dealers_to_show = df_unarchive_dealers_to_show.rename(
        columns=rename_dict
    )
    output_4 = df_unarchive_dealers_to_show.copy()
    print(tabulate(output_4, headers="keys", tablefmt="pretty", showindex=False))
    print()
    print("*" * 100)

    return (
        title_1_str,
        title_2_str,
        output_2,
        title_3_str,
        output_3,
        title_4_str,
        output_4,
    )


def show_dealer_results_main(
    df_dealer_dealer_results,
    df_dealer_total_scanning_locations,
    df_dealer_total_centroids,
    dealer_scope_dict_path,
):

    if df_dealer_total_scanning_locations.empty:
        print("无记录")
        return

    dealer_info_cols = [
        "is_dealer_suspicious",
        "is_dealer_no_valid_scope",
        "dealer_suspicious_points_count",
        "dealer_remote_scanning_count",
        "dealer_suspicious_points_ratio",
        "dealer_remote_ratio",
        "dealer_total_scanning_count",
        "dealer_total_box_count",
        "dealer_suspicious_hotspot_count",
        "dealer_suspicious_hotspot_ratio",
        "dealer_hotspot_count",
    ]

    dealer_cluster_cols = [
        "cluster_label",
        "province",
        "city",
        "district",
        "street",
        "scanning_count_within_cluster",
        "is_remote",
        "is_suspicious",
        "scanning_ratio_for_cluster",
        "box_count_within_cluster",
        "dis_to_all_local_hotspots_centroid",
        "dis_to_all_local_points_centroid",
        "dis_border",
        "std_distance_within_cluster",
    ]

    rename_dict = {
        "is_dealer_suspicious": "是否可疑",
        "is_dealer_no_valid_scope": "无<有效>经营范围",
        "dealer_suspicious_points_count": "可疑扫码数量",
        "dealer_remote_scanning_count": "异地扫码数量",
        "dealer_suspicious_points_ratio": "扫码可疑率",
        "dealer_remote_ratio": "扫码异地率",
        "dealer_total_scanning_count": "总扫码量",
        "dealer_total_box_count": "总箱数",
        "dealer_suspicious_hotspot_count": "可疑扫码热点数量",
        "dealer_remote_hotspot_count": "异地扫码热点数量",
        "dealer_suspicious_hotspot_ratio": "热点可疑率",
        "dealer_remote_hotspot_ratio": "热点异地率",
        "dealer_hotspot_count": "热点总数量",
        "cluster_label": "簇标签",
        "province": "省",
        "city": "市",
        "district": "区",
        "street": "镇街",
        "scanning_count_within_cluster": "扫码数",
        "is_remote": "异地",
        "is_suspicious": "高度可疑",
        "scanning_ratio_for_cluster": "扫码量占比",
        "dis_to_all_local_hotspots_centroid": "距本地热点总质心",
        "dis_to_all_local_points_centroid": "距本地点总质心",
        "std_distance_within_cluster": "离散度",
        "box_count_within_cluster": "箱数",
        "dis_border": "距边界",
        "BELONG_DEALER_NO": "经销商编码",
        "BELONG_DEALER_NAME": "经销商名称",
        "PRODUCT_GROUP_NAME": "品项名称",
        "EFFECTIVE_DATE": "生效日期",
        "INACTIVE_DATE": "失效日期",
        "DEALER_CODE": "经销商编码",
        "PRODUCT_GROUP_CODE": "品项编码",
        "AREA_NAME": "经营区域",
        "AREA_CODE": "区域代码",
        "PROVINCE": "省",
        "CITY": "市",
        "DISTRICT": "区县",
        "STREET": "镇街",
        "OPEN_PROVINCE": "省",
        "OPEN_CITY": "市",
        "count": "数量",
        "is_remote_city": "是否为异地二级城市",
    }

    dealer_id = df_dealer_total_scanning_locations.loc[0, "BELONG_DEALER_NO"]
    dealer_name = df_dealer_total_scanning_locations.loc[0, "BELONG_DEALER_NAME"]
    product_group_id = df_dealer_total_scanning_locations.loc[0, "PRODUCT_GROUP_CODE"]
    product_group_name = df_dealer_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]

    # Output_1
    title_1_str = f"经销商: {dealer_id} - {dealer_name} - {product_group_name}"
    print(title_1_str)
    print("-" * 60)
    print()

    # Output_2
    title_2_str = "--- 经营范围 ---"
    print(title_2_str)
    with open(dealer_scope_dict_path, "rb") as f:
        dealer_scope_dict = pickle.load(f)

    df_business_scope = dealer_scope_dict.get((dealer_id, product_group_id), None)
    if (dealer_id, product_group_id) not in dealer_scope_dict:
        print(
            f"{dealer_id} - {dealer_name} - {product_group_name}的经营范围并未记录在档! "
        )
        return

    df_business_scope = dealer_scope_dict[(dealer_id, product_group_id)]
    df_business_scope_to_show = df_business_scope.copy()
    df_business_scope_to_show["EFFECTIVE_DATE"] = df_business_scope_to_show[
        "EFFECTIVE_DATE"
    ].dt.strftime("%Y-%m-%d")
    df_business_scope_to_show["INACTIVE_DATE"] = df_business_scope_to_show[
        "INACTIVE_DATE"
    ].dt.strftime("%Y-%m-%d")
    # df_business_scope_to_show = df_business_scope_to_show.drop(
    #     columns="AREA_CODE"
    # )
    df_business_scope_to_show = df_business_scope_to_show.rename(columns=rename_dict)

    output_2 = df_business_scope_to_show.copy()
    print(
        tabulate(
            df_business_scope_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_3
    title_3_str = "--- 范围内经销商异地信息 ---"
    print(title_3_str)
    df_dealer_info_to_show = df_dealer_dealer_results.loc[:, dealer_info_cols]
    df_dealer_info_to_show[["is_dealer_suspicious", "is_dealer_no_valid_scope"]] = (
        df_dealer_info_to_show[
            ["is_dealer_suspicious", "is_dealer_no_valid_scope"]
        ].replace({1: "是", 0: "否"})
    )
    df_dealer_info_to_show[
        [
            "dealer_suspicious_points_ratio",
            "dealer_remote_ratio",
            "dealer_suspicious_hotspot_ratio",
        ]
    ] = df_dealer_info_to_show[
        [
            "dealer_suspicious_points_ratio",
            "dealer_remote_ratio",
            "dealer_suspicious_hotspot_ratio",
        ]
    ].round(
        2
    )
    df_dealer_info_to_show = df_dealer_info_to_show.rename(columns=rename_dict)
    output_3 = df_dealer_info_to_show.copy()
    print(
        tabulate(
            df_dealer_info_to_show, headers="keys", tablefmt="pretty", showindex=False
        )
    )
    print()

    # Output_4
    df_dealer_cluster = df_dealer_total_centroids.loc[
        df_dealer_total_centroids.dealer_id == dealer_id, dealer_cluster_cols
    ]
    df_dealer_cluster = df_dealer_cluster.loc[
        ~(df_dealer_cluster["cluster_label"].isin([-2, -1])), :
    ]
    suspicious_labels = list(
        map(
            int,
            list(
                df_dealer_cluster.loc[
                    df_dealer_cluster["is_suspicious"] == 1, "cluster_label"
                ].values
            ),
        )
    )
    df_dealer_cluster_to_show = df_dealer_cluster.sort_values(
        by=["is_suspicious", "cluster_label"], ascending=[False, True]
    )
    df_dealer_cluster_to_show[
        ["scanning_count_within_cluster", "box_count_within_cluster"]
    ] = df_dealer_cluster_to_show[
        ["scanning_count_within_cluster", "box_count_within_cluster"]
    ].astype(
        int
    )
    df_dealer_cluster_to_show["scanning_ratio_for_cluster"] = df_dealer_cluster_to_show[
        "scanning_ratio_for_cluster"
    ].round(2)
    df_dealer_cluster_to_show[["is_remote", "is_suspicious"]] = (
        df_dealer_cluster_to_show[["is_remote", "is_suspicious"]].replace(
            {1: "是", 0: "否"}
        )
    )
    df_dealer_cluster_to_show = df_dealer_cluster_to_show.rename(columns=rename_dict)

    title_4_str = f"--- 经销商热点信息 ---\n可疑热点标签:\n{str(suspicious_labels)}"
    print(title_4_str)
    output_4 = df_dealer_cluster_to_show.copy()
    print(
        tabulate(
            df_dealer_cluster_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_5
    df_city_count = (
        df_dealer_total_scanning_locations[["OPEN_PROVINCE", "OPEN_CITY"]]
        .value_counts()
        .reset_index()
    )
    # 扫码表格式 -> 经营范围表格式
    # ['天津市', '天津市'] -> ['天津', '天津市']
    df_city_count["OPEN_PROVINCE"] = df_city_count["OPEN_PROVINCE"].apply(
        lambda x: x[:2] if x in ["天津市", "北京市", "上海市", "重庆市"] else x
    )
    df_valid_scope = df_dealer_dealer_results.loc[0, "dealer_valid_scope"]
    df_count_merged = pd.DataFrame()

    if df_valid_scope.empty:
        df_count_merged = (
            df_city_count.loc[:, ["OPEN_PROVINCE", "OPEN_CITY", "count"]]
            .drop_duplicates()
            .reset_index(drop=True)
        )
        df_count_merged["is_remote_city"] = "是"

    else:
        df_count_merged = df_city_count.merge(
            df_valid_scope,
            left_on=["OPEN_PROVINCE", "OPEN_CITY"],
            right_on=["PROVINCE", "CITY"],
            how="left",
            indicator=True,
        )
        df_count_merged["is_remote_city"] = df_count_merged["_merge"] == "left_only"
        df_count_merged = (
            df_count_merged.loc[
                :, ["OPEN_PROVINCE", "OPEN_CITY", "count", "is_remote_city"]
            ]
            .drop_duplicates()
            .reset_index(drop=True)
        )
        df_count_merged["is_remote_city"] = df_count_merged["is_remote_city"].map(
            {True: "是", False: "否"}
        )

    df_count_merged_to_show = df_count_merged.loc[df_count_merged["count"] > 5].rename(
        columns=rename_dict
    )

    title_5_str = f"--- 可疑经销商开瓶城市统计表(大于五瓶的城市）--- \n开瓶二级城市数: {len(df_city_count)}"
    print(title_5_str)
    output_5 = df_count_merged_to_show.copy()
    print(
        tabulate(
            df_count_merged_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_6
    title_6_str = "---热点地图---"
    polyline_points_list_total = df_dealer_dealer_results.loc[
        0, "dealer_polyline_points_list_total"
    ]
    # adcodes = df_dealer_dealer_results.loc[0, 'dealer_adcodes']
    m = plot_clusters_with_folium(
        df_dealer_total_scanning_locations,
        points_size=3,
        noise_size=1,
        polyline_points_list=polyline_points_list_total,
    )
    output_6 = m
    display(m)
    print()
    print()
    print()

    return (
        title_1_str,
        title_2_str,
        output_2,
        title_3_str,
        output_3,
        title_4_str,
        output_4,
        title_5_str,
        output_5,
        title_6_str,
        output_6,
    )


def show_dealer_results_special_main(
    df_dealer_dealer_results,
    df_dealer_total_scanning_locations,
    df_dealer_total_centroids,
    df_dealer_dealer_results_dense,
    df_dealer_total_scanning_locations_dense,
    df_dealer_total_centroids_dense,
    dealer_scope_dict_path,
):

    if df_dealer_total_scanning_locations.empty:
        print("无记录")
        return

    dealer_info_cols = [
        "is_dealer_suspicious_final",
        "is_dealer_no_valid_scope",
        "dealer_suspicious_points_count_final",
        "dealer_remote_scanning_count",
        "dealer_suspicious_points_ratio_final",
        "dealer_remote_ratio",
        "dealer_total_scanning_count",
        "dealer_total_box_count",
        "dealer_suspicious_hotspot_count_final",
        "dealer_suspicious_hotspot_ratio_final",
        "dealer_hotspot_count_final",
    ]

    dealer_cluster_cols = [
        "cluster_label",
        "province",
        "city",
        "district",
        "street",
        "scanning_count_within_cluster",
        "is_remote",
        "is_suspicious",
        "scanning_ratio_for_cluster",
        "box_count_within_cluster",
        "dis_to_all_local_hotspots_centroid",
        "dis_to_all_local_points_centroid",
        "dis_border",
        "std_distance_within_cluster",
    ]

    rename_dict = {
        "EFFECTIVE_DATE": "生效日期",
        "INACTIVE_DATE": "失效日期",
        "DEALER_CODE": "经销商编码",
        "PRODUCT_GROUP_CODE": "品项编码",
        "AREA_NAME": "经营区域",
        "AREA_CODE": "区域代码",
        "PROVINCE": "省",
        "CITY": "市",
        "DISTRICT": "区县",
        "STREET": "镇街",
        "is_dealer_suspicious_final": "是否可疑",
        "is_dealer_no_valid_scope": "无<有效>经营范围",
        "dealer_suspicious_points_count_final": "可疑扫码数量",
        "dealer_remote_scanning_count": "异地扫码数量",
        "dealer_suspicious_points_ratio_final": "扫码可疑率",
        "dealer_remote_ratio": "扫码异地率",
        "dealer_total_scanning_count": "总扫码量",
        "dealer_total_box_count": "总箱数",
        "dealer_suspicious_hotspot_count_final": "可疑扫码热点数量",
        "dealer_suspicious_hotspot_ratio_final": "热点可疑率",
        "dealer_hotspot_count_final": "热点总数量",
        "dealer_hotspot_count_from_sparse": "一级热点数",
        "dealer_suspicious_hotspot_count_from_sparse": "一级可疑热点数",
        "dealer_remote_hotspot_count_sparse": "一级异地热点数",
        "dealer_suspicious_hotspot_ratio_sparse": "一级热点可疑率",
        "dealer_remote_hotspot_ratio_sparse": "一级热点异地率",
        "dealer_hotspot_count_dense": "二级热点数",
        "dealer_suspicious_hotspot_count_dense": "二级可疑热点数",
        "dealer_remote_hotspot_count_dense": "二级异地热点数",
        "dealer_suspicious_hotspot_ratio_dense": "二级热点可疑率",
        "dealer_remote_hotspot_ratio_dense": "二级热点异地率",
        "OPEN_PROVINCE": "省",
        "OPEN_CITY": "市",
        "count": "数量",
        "is_remote_city": "是否为异地二级城市",
        "cluster_label": "簇标签",
        "province": "省",
        "city": "市",
        "district": "区",
        "street": "镇街",
        "scanning_count_within_cluster": "扫码数",
        "is_remote": "异地",
        "is_suspicious": "高度可疑",
        "scanning_ratio_for_cluster": "扫码量占比",
        "dis_to_all_local_hotspots_centroid": "距本地热点总质心",
        "dis_to_all_local_points_centroid": "距本地点总质心",
        "std_distance_within_cluster": "离散度",
        "box_count_within_cluster": "箱数",
        "dis_border": "距边界",
    }

    dealer_id = df_dealer_total_scanning_locations.loc[0, "BELONG_DEALER_NO"]
    dealer_name = df_dealer_total_scanning_locations.loc[0, "BELONG_DEALER_NAME"]
    product_group_id = df_dealer_total_scanning_locations.loc[0, "PRODUCT_GROUP_CODE"]
    product_group_name = df_dealer_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]

    df_dealer_hotspots = df_dealer_total_centroids.loc[
        ~(df_dealer_total_centroids["cluster_label"].isin([-2, -1])), :
    ]
    df_dealer_hotspots_dense = df_dealer_total_centroids_dense.loc[
        ~(df_dealer_total_centroids_dense["cluster_label"].isin([-2, -1])), :
    ]

    # Output_1
    title_1_str = f"经销商: {dealer_id} - {dealer_name} - {product_group_name}"
    print(title_1_str)
    print("-" * 60)
    print()

    # Output_2
    title_2_str = "---经营范围---"
    print(title_2_str)
    with open(dealer_scope_dict_path, "rb") as f:
        dealer_scope_dict = pickle.load(f)

    df_business_scope = dealer_scope_dict.get((dealer_id, product_group_id), None)
    if (dealer_id, product_group_id) not in dealer_scope_dict:
        print(
            f"{dealer_id} - {dealer_name} - {product_group_name}的经营范围并未记录在档! "
        )
        return
    df_business_scope_to_show = (
        df_business_scope.copy()
    )  # 不知道为什么会有bug:会更改到字典里的数据类型,因此采用深拷贝
    df_business_scope_to_show["EFFECTIVE_DATE"] = df_business_scope_to_show[
        "EFFECTIVE_DATE"
    ].dt.strftime("%Y-%m-%d")
    df_business_scope_to_show["INACTIVE_DATE"] = df_business_scope_to_show[
        "INACTIVE_DATE"
    ].dt.strftime("%Y-%m-%d")
    # df_business_scope_to_show = df_business_scope_to_show.drop(
    #     columns="AREA_CODE"
    # )
    df_business_scope_to_show = df_business_scope_to_show.rename(columns=rename_dict)

    output_2 = df_business_scope_to_show.copy()
    print(
        tabulate(
            df_business_scope_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_3
    title_3_str = "---范围内经销商异地信息---"
    print(title_3_str)
    df_dealer_info_to_show = df_dealer_dealer_results.loc[:, dealer_info_cols]
    df_dealer_info_to_show[
        ["is_dealer_suspicious_final", "is_dealer_no_valid_scope"]
    ] = df_dealer_info_to_show[
        ["is_dealer_suspicious_final", "is_dealer_no_valid_scope"]
    ].replace(
        {1: "是", 0: "否"}
    )
    df_dealer_info_to_show[
        [
            "dealer_suspicious_points_ratio_final",
            "dealer_remote_ratio",
            "dealer_suspicious_hotspot_ratio_final",
        ]
    ] = df_dealer_info_to_show[
        [
            "dealer_suspicious_points_ratio_final",
            "dealer_remote_ratio",
            "dealer_suspicious_hotspot_ratio_final",
        ]
    ].round(
        2
    )
    df_dealer_info_to_show = df_dealer_info_to_show.rename(columns=rename_dict)

    output_3 = df_dealer_info_to_show
    print(
        tabulate(
            df_dealer_info_to_show, headers="keys", tablefmt="pretty", showindex=False
        )
    )
    print()

    # Output_4
    title_4_str = "---一级热点信息---"
    print(title_4_str)
    dealer_info_cols_sparse = [
        "dealer_hotspot_count_from_sparse",
        "dealer_suspicious_hotspot_count_from_sparse",
        "dealer_remote_hotspot_count",
        "dealer_suspicious_hotspot_ratio",
        "dealer_remote_hotspot_ratio",
    ]
    dealer_info_cols_dense = [
        "dealer_hotspot_count",
        "dealer_suspicious_hotspot_count",
        "dealer_remote_hotspot_count",
        "dealer_suspicious_hotspot_ratio",
        "dealer_remote_hotspot_ratio",
    ]
    sparse_info_cols_dict = {
        "dealer_remote_hotspot_count": "dealer_remote_hotspot_count_sparse",
        "dealer_suspicious_hotspot_ratio": "dealer_suspicious_hotspot_ratio_sparse",
        "dealer_remote_hotspot_ratio": "dealer_remote_hotspot_ratio_sparse",
    }
    dense_info_cols_dict = {
        "dealer_hotspot_count": "dealer_hotspot_count_dense",
        "dealer_suspicious_hotspot_count": "dealer_suspicious_hotspot_count_dense",
        "dealer_remote_hotspot_count": "dealer_remote_hotspot_count_dense",
        "dealer_suspicious_hotspot_ratio": "dealer_suspicious_hotspot_ratio_dense",
        "dealer_remote_hotspot_ratio": "dealer_remote_hotspot_ratio_dense",
    }

    df_dealer_info_sparse_to_show = df_dealer_dealer_results.loc[
        :, dealer_info_cols_sparse
    ].rename(columns=sparse_info_cols_dict)
    df_dealer_info_sparse_to_show[
        [
            "dealer_hotspot_count_from_sparse",
            "dealer_suspicious_hotspot_count_from_sparse",
            "dealer_remote_hotspot_count_sparse",
        ]
    ] = df_dealer_info_sparse_to_show[
        [
            "dealer_hotspot_count_from_sparse",
            "dealer_suspicious_hotspot_count_from_sparse",
            "dealer_remote_hotspot_count_sparse",
        ]
    ].astype(
        int
    )
    df_dealer_info_sparse_to_show[
        ["dealer_suspicious_hotspot_ratio_sparse", "dealer_remote_hotspot_ratio_sparse"]
    ] = df_dealer_info_sparse_to_show[
        ["dealer_suspicious_hotspot_ratio_sparse", "dealer_remote_hotspot_ratio_sparse"]
    ].round(
        2
    )
    df_dealer_info_sparse_to_show = df_dealer_info_sparse_to_show.rename(
        columns=rename_dict
    )

    output_4 = df_dealer_info_sparse_to_show.copy()
    print(
        tabulate(
            df_dealer_info_sparse_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )

    # Output_5
    title_5_str = "---二级级热点信息---"
    output_5 = []
    dense_empty = df_dealer_dealer_results_dense.empty

    if not dense_empty:
        df_dealer_info_dense_to_show = df_dealer_dealer_results_dense.loc[
            :, dealer_info_cols_dense
        ].rename(columns=dense_info_cols_dict)
        df_dealer_info_dense_to_show[
            [
                "dealer_hotspot_count_dense",
                "dealer_suspicious_hotspot_count_dense",
                "dealer_remote_hotspot_count_dense",
            ]
        ] = df_dealer_info_dense_to_show[
            [
                "dealer_hotspot_count_dense",
                "dealer_suspicious_hotspot_count_dense",
                "dealer_remote_hotspot_count_dense",
            ]
        ].astype(
            int
        )

        df_dealer_info_dense_to_show[
            [
                "dealer_suspicious_hotspot_ratio_dense",
                "dealer_remote_hotspot_ratio_dense",
            ]
        ] = df_dealer_info_dense_to_show[
            [
                "dealer_suspicious_hotspot_ratio_dense",
                "dealer_remote_hotspot_ratio_dense",
            ]
        ].round(
            2
        )
        # df_dealer_info_dense_to_show[
        #     "dealer_hotspot_count_dense"
        # ] -= df_dealer_info_dense_to_show["dealer_hotspot_count_dense"].astype(int)
        df_dealer_info_dense_to_show = df_dealer_info_dense_to_show.rename(
            columns=rename_dict
        )

    else:
        cols = [
            "dealer_hotspot_count_dense",
            "dealer_suspicious_hotspot_count_dense",
            "dealer_remote_hotspot_count_dense",
            "dealer_suspicious_hotspot_ratio_dense",
            "dealer_remote_hotspot_ratio_dense",
        ]
        df_dealer_info_dense_to_show = pd.DataFrame(0, index=[0], columns=cols).rename(
            columns=rename_dict
        )

    output_5 = df_dealer_info_dense_to_show.copy()
    print(title_5_str)
    print(
        tabulate(
            df_dealer_info_dense_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_6
    # 打印开瓶二级城市统计
    df_city_count = (
        df_dealer_total_scanning_locations[["OPEN_PROVINCE", "OPEN_CITY"]]
        .value_counts()
        .reset_index()
    )
    # 扫码表格式 -> 经营范围表格式
    # ['天津市', '天津市'] -> ['天津', '天津市']
    df_city_count["OPEN_PROVINCE"] = df_city_count["OPEN_PROVINCE"].apply(
        lambda x: x[:2] if x in ["天津市", "北京市", "上海市", "重庆市"] else x
    )

    df_valid_scope = df_dealer_dealer_results.loc[0, "dealer_valid_scope"]
    df_count_merged = pd.DataFrame()

    if df_valid_scope.empty:
        df_count_merged = (
            df_city_count.loc[:, ["OPEN_PROVINCE", "OPEN_CITY", "count"]]
            .drop_duplicates()
            .reset_index(drop=True)
        )
        df_count_merged["is_remote_city"] = "是"

    else:
        df_count_merged = df_city_count.merge(
            df_valid_scope,
            left_on=["OPEN_PROVINCE", "OPEN_CITY"],
            right_on=["PROVINCE", "CITY"],
            how="left",
            indicator=True,
        )
        df_count_merged["is_remote_city"] = df_count_merged["_merge"] == "left_only"
        df_count_merged = (
            df_count_merged.loc[
                :, ["OPEN_PROVINCE", "OPEN_CITY", "count", "is_remote_city"]
            ]
            .drop_duplicates()
            .reset_index(drop=True)
        )
        df_count_merged["is_remote_city"] = df_count_merged["is_remote_city"].map(
            {True: "是", False: "否"}
        )

    df_count_merged_to_show = df_count_merged.loc[df_count_merged["count"] > 5].rename(
        columns=rename_dict
    )

    title_6_str = f"--- 可疑经销商开瓶城市统计表(大于五瓶的城市）--- \n开瓶二级城市数: {len(df_city_count)}"
    print(title_6_str)

    output_6 = df_count_merged_to_show.copy()
    print(
        tabulate(
            df_count_merged_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_7
    polyline_points_list_total = df_dealer_dealer_results.loc[
        0, "dealer_polyline_points_list_total"
    ]

    df_dealer_cluster = df_dealer_hotspots.loc[:, dealer_cluster_cols]
    suspicious_labels = list(
        map(
            int,
            list(
                df_dealer_cluster.loc[
                    df_dealer_cluster["is_suspicious"] == 1, "cluster_label"
                ].values
            ),
        )
    )
    title_7_str = f"---经销商热点信息---\n可疑一级热点标签:\n{str(suspicious_labels)}"
    print(title_7_str)
    df_dealer_cluster_to_show = df_dealer_cluster.sort_values(
        by=["is_suspicious", "cluster_label"], ascending=[False, True]
    )
    df_dealer_cluster_to_show[
        ["scanning_count_within_cluster", "box_count_within_cluster"]
    ] = df_dealer_cluster_to_show[
        ["scanning_count_within_cluster", "box_count_within_cluster"]
    ].astype(
        int
    )
    df_dealer_cluster_to_show["scanning_ratio_for_cluster"] = df_dealer_cluster_to_show[
        "scanning_ratio_for_cluster"
    ].round(3)
    df_dealer_cluster_to_show[["is_remote", "is_suspicious"]] = (
        df_dealer_cluster_to_show[["is_remote", "is_suspicious"]].replace(
            {1: "是", 0: "否"}
        )
    )
    df_dealer_cluster_to_show = df_dealer_cluster_to_show.rename(columns=rename_dict)

    output_7 = df_dealer_cluster_to_show.copy()
    print(
        tabulate(
            df_dealer_cluster_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_8
    title_8_str = "---一级热点地图---"
    print(title_8_str)
    m = plot_clusters_with_folium(
        df_dealer_total_scanning_locations,
        points_size=3,
        noise_size=1,
        polyline_points_list=polyline_points_list_total,
    )
    output_8 = m
    display(m)

    # Output_9
    df_dealer_cluster_dense = df_dealer_hotspots_dense.loc[:, dealer_cluster_cols]
    suspicious_labels_dense = list(
        map(
            int,
            list(
                df_dealer_cluster_dense.loc[
                    df_dealer_cluster_dense["is_suspicious"] == 1, "cluster_label"
                ].values
            ),
        )
    )
    title_9_str = f"可疑二级热点标签:\n{str(suspicious_labels_dense)}"
    print(title_9_str)

    df_dealer_cluster_dense_to_show = df_dealer_cluster_dense.sort_values(
        by=["is_suspicious", "cluster_label"], ascending=[False, True]
    )
    df_dealer_cluster_dense_to_show[
        ["scanning_count_within_cluster", "box_count_within_cluster"]
    ] = df_dealer_cluster_dense_to_show[
        ["scanning_count_within_cluster", "box_count_within_cluster"]
    ].astype(
        int
    )
    df_dealer_cluster_dense_to_show["scanning_ratio_for_cluster"] = (
        df_dealer_cluster_dense_to_show["scanning_ratio_for_cluster"].round(3)
    )
    df_dealer_cluster_dense_to_show[["is_remote", "is_suspicious"]] = (
        df_dealer_cluster_dense_to_show[["is_remote", "is_suspicious"]].replace(
            {1: "是", 0: "否"}
        )
    )
    df_dealer_cluster_dense_to_show = df_dealer_cluster_dense_to_show.rename(
        columns=rename_dict
    )
    output_9 = df_dealer_cluster_dense_to_show.copy()
    print(
        tabulate(
            df_dealer_cluster_dense_to_show,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    # Output_10
    title_10_str = "---二级热点地图---"
    print(title_10_str)
    if not dense_empty:
        m2 = plot_clusters_with_folium(
            df_dealer_total_scanning_locations_dense,
            points_size=3,
            noise_size=1,
            polyline_points_list=polyline_points_list_total,
        )
        output_10 = m2
        display(m2)
    else:
        output_10 = "---无二级热点---"
        print(output_10)
    print()
    print()
    print()
    return (
        title_1_str,
        title_2_str,
        output_2,
        title_3_str,
        output_3,
        title_4_str,
        output_4,
        title_5_str,
        output_5,
        title_6_str,
        output_6,
        title_7_str,
        output_7,
        title_8_str,
        output_8,  # map1
        title_9_str,
        output_9,
        title_10_str,
        output_10,  # map2
    )


def generate_single_dealer_sheet(
    results_files_folder_path,
    excel_file_name_str,
    map_file_name_str,
    sheet_name,
    title_1_str,
    title_2_str,
    output_2,
    title_3_str,
    output_3,
    title_4_str,
    output_4,
    title_5_str,
    output_5,
    title_6_str,
    output_6,
    mode,
    **if_sheet_exists,
):

    excel_results_path = os.path.join(
        results_files_folder_path,
        f"{excel_file_name_str}.xlsx",
    )

    with pd.ExcelWriter(
        excel_results_path, engine="openpyxl", mode=mode, **if_sheet_exists
    ) as writer:
        if mode == "w":
            sheet1 = writer.book.create_sheet(sheet_name)
        if mode == "a":
            sheet1 = writer.book[sheet_name]
        blue_fill = PatternFill(
            start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
        )  # 浅蓝色
        yellow_fill = PatternFill(
            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
        )
        soft_red_fill = PatternFill(
            start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
        )
        # 写入标题 1
        title_1_start_row = 2
        title_1_start_col = 1
        cell = sheet1.cell(
            row=title_1_start_row, column=title_1_start_col, value=title_1_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=14)
        # 判断是否因 无有效经营范围 引发的可疑
        if output_3.loc[0, "无<有效>经营范围"] == "否":
            cell.fill = soft_red_fill
        else:
            cell.fill = yellow_fill
        sheet1.merge_cells(
            start_row=title_1_start_row,
            end_row=title_1_start_row,
            start_column=title_1_start_col,
            end_column=title_1_start_col + 10,
        )

        # 写入标题 2 和 output_2
        title_2_start_row = 4
        title_2_start_col = 1
        output_2_start_row = title_2_start_row
        cell = sheet1.cell(
            row=title_2_start_row, column=title_2_start_col, value=title_2_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_2_start_row,
            end_row=title_2_start_row,
            start_column=title_2_start_col,
            end_column=title_2_start_col + (output_2.shape[1]) - 1,
        )
        output_2.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_2_start_row
        )

        title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
        title_3_start_col = 1
        output_3_start_row = title_3_start_row
        cell = sheet1.cell(
            row=title_3_start_row, column=title_3_start_col, value=title_3_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_3_start_row,
            end_row=title_3_start_row,
            start_column=title_3_start_col,
            end_column=title_3_start_col + output_3.shape[1] - 1,
        )
        output_3.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_3_start_row
        )

        # 4 具体簇的信息
        title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
        title_4_start_col = 1
        output_4_start_row = title_4_start_row
        cell = sheet1.cell(
            row=title_4_start_row, column=title_4_start_col, value=title_4_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = blue_fill
        cell.font = Font(bold=True, size=12)
        sheet1.merge_cells(
            start_row=title_4_start_row,
            end_row=title_4_start_row,
            start_column=title_4_start_col,
            end_column=title_4_start_col + output_4.shape[1] - 1,
        )
        # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
        output_4[["距本地热点总质心", "距本地点总质心"]] = output_4[
            ["距本地热点总质心", "距本地点总质心"]
        ].fillna("na")
        output_4.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_4_start_row
        )
        df_suspicious = output_4.loc[output_4["高度可疑"] == "是", :]
        for row_no in range(
            output_4_start_row + 2,
            output_4_start_row + df_suspicious.shape[0] + 2,
        ):
            for col_no in range(
                title_4_start_col, title_4_start_col + df_suspicious.shape[1]
            ):
                cell = sheet1.cell(row=row_no, column=col_no)
                cell.fill = soft_red_fill

        # 5
        title_5_start_row = output_4_start_row + len(output_4) + 1 + 2
        title_5_start_col = 1
        output_5_start_row = title_5_start_row
        cell = sheet1.cell(
            row=title_5_start_row, column=title_5_start_col, value=title_5_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_5_start_row,
            end_row=title_5_start_row,
            start_column=title_5_start_col,
            end_column=title_5_start_col + output_5.shape[1] - 1,
        )
        output_5.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_5_start_row
        )

        # 6 output是map
        map_results_path = os.path.join(
            results_files_folder_path,
            f"{map_file_name_str}.html",
        )
        output_6.save(map_results_path)

    # 打开已保存的 Excel 文件，进行进一步样式设置
    wb = load_workbook(excel_results_path)
    ws = wb[sheet_name]

    # 设置列宽
    for i in range(14):
        ascii_value = ord("A")
        col_letter = chr(ascii_value + i)
        ws.column_dimensions[col_letter].width = 20
    # ws.column_dimensions['A'].width = 30
    # ws.column_dimensions['B'].width = 40
    # ws.column_dimensions['C'].width = 30
    # 遍历每个单元格并设置居中对齐
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(excel_results_path)


def show_results_main(
    df_dealer_results,
    df_total_scanning_locations,
    df_total_centroids,
    df_suspicious_hotspots_parameters,
    dealer_scope_dict_path,
    dealer_region_name,
    product_group_id,
    year_month_str,
    save_results=True,
):
    start_date_str, end_date_str = get_month_start_end(year_month_str)

    results_files_folder_path = (
        f"results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
    )
    os.makedirs(results_files_folder_path, exist_ok=True)

    rename_dict = {
        "radius": "簇半径",
        "min_samples": "簇内最少样本数",
        "dis_hotspots_c_t": "距本地热点总质心的距离阈值",
        "dis_points_c_t": "距本地扫码点总质心的距离阈值",
        "dis_border_t": "距边界最小距离阈值",
        "ratio_scanning_t": "热点扫码量占比阈值",
        "scanning_count_t": "热点扫码量阈值",
        "std_distance_t": "热点离散度阈值",
        "box_count_t": "紧密热点的箱数阈值",
        "BELONG_DEALER_NO": "经销商编码",
        "BELONG_DEALER_NAME": "经销商名称",
        "PRODUCT_GROUP_NAME": "品项名称",
    }

    df_model_parameters = df_suspicious_hotspots_parameters.copy().rename(
        columns=rename_dict
    )
    df_suspicious_dealers = df_dealer_results.loc[
        df_dealer_results.is_dealer_suspicious == 1, :
    ]
    df_suspicious_dealers = df_suspicious_dealers.sort_values(
        by="BELONG_DEALER_NO"
    ).reset_index(drop=True)

    product_group_name = df_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]
    main_title_str_1 = (
        "-" * 35
        + f"基于模型v1.0, ({start_date_str} - {end_date_str}),  ({dealer_region_name} - {product_group_name})的结果如下"
        + "-" * 35
    )
    print(main_title_str_1)
    print()

    main_title_str_2 = "--- 模型参数 ---"
    print(main_title_str_2)
    print(
        tabulate(
            df_model_parameters, headers="keys", tablefmt="pretty", showindex=False
        )
    )
    print()

    # show_region_short_results-----------------------
    (
        title_1_str,
        title_2_str,
        output_2,
        title_3_str,
        output_3,
        title_4_str,
        output_4,
    ) = show_region_short_results(
        df_dealer_results,
        df_total_scanning_locations,
        start_date_str,
        end_date_str,
        dealer_region_name,
    )

    # save excel part
    if save_results:
        if os.path.exists(results_files_folder_path):
            # Use glob to find all files in the directory
            files = glob.glob(
                os.path.join(results_files_folder_path, "*")
            )  # This matches all files and subdirectories

            for file in files:
                # Check if it's a file (not a subdirectory)
                if os.path.isfile(file):
                    os.remove(file)
        file_name_str = f"{dealer_region_name}-{year_month_str}-{product_group_name}"
        excel_results_path = os.path.join(
            results_files_folder_path,
            f"{file_name_str}.xlsx",
        )

        # 输出大区汇总excel
        with pd.ExcelWriter(excel_results_path, engine="openpyxl") as writer:
            # 创建工作表

            sheet1 = writer.book.create_sheet("大区汇总信息")
            blue_fill = PatternFill(
                start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
            )  # 浅蓝色
            yellow_fill = PatternFill(
                start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
            )
            soft_red_fill = PatternFill(
                start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
            )
            hyperlink_font = Font(color="0000FF", underline="single")
            ids = list(output_3["经销商编码"])
            for i, id in enumerate(ids):
                writer.book.create_sheet(id)

            # 写入标题 1
            title_1_start_row = 1
            title_1_start_col = 1
            cell = sheet1.cell(
                row=title_1_start_row, column=title_1_start_col, value=title_1_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=14)
            sheet1.merge_cells(
                start_row=title_1_start_row,
                end_row=title_1_start_row,
                start_column=title_1_start_col,
                end_column=title_1_start_col + 11,
            )

            # 写入标题 2 和 output_2
            title_2_start_row = 3
            title_2_start_col = 1
            output_2_start_row = title_2_start_row
            cell = sheet1.cell(
                row=title_2_start_row, column=title_2_start_col, value=title_2_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=12)
            cell.fill = blue_fill
            sheet1.merge_cells(
                start_row=title_2_start_row,
                end_row=title_2_start_row,
                start_column=title_2_start_col,
                end_column=title_2_start_col + output_2.shape[1] - 1,
            )
            output_2.to_excel(
                writer,
                sheet_name="大区汇总信息",
                index=False,
                startrow=output_2_start_row,
            )

            title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
            title_3_start_col = 1
            output_3_start_row = title_3_start_row
            cell = sheet1.cell(
                row=title_3_start_row, column=title_3_start_col, value=title_3_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=12)
            cell.fill = blue_fill
            sheet1.merge_cells(
                start_row=title_3_start_row,
                end_row=title_3_start_row,
                start_column=title_3_start_col,
                end_column=title_3_start_col + output_3.shape[1] - 1,
            )
            output_3.to_excel(
                writer,
                sheet_name="大区汇总信息",
                index=False,
                startrow=output_3_start_row,
            )

            i = 0
            for row_no in range(
                output_3_start_row + 2, output_3_start_row + len(output_3) + 2
            ):
                cell_hyper = sheet1.cell(row=row_no, column=1)
                cell_hyper.hyperlink = f"#{ids[i]}!A1"
                cell_hyper.font = hyperlink_font
                i += 1
            # 并非 无任何有效经营范围而引发的可疑 填充红色
            df_no_valid_region = output_3.loc[output_3["无<有效>经营范围"] == "否", :]
            for row_no in range(
                output_3_start_row + 2,
                output_3_start_row + df_no_valid_region.shape[0] + 2,
            ):
                for col_no in range(
                    title_3_start_col, title_3_start_col + df_no_valid_region.shape[1]
                ):
                    cell = sheet1.cell(row=row_no, column=col_no)
                    cell.fill = soft_red_fill
            # 无任何有效经营范围引发的可疑 填充黄色
            for row_no in range(
                output_3_start_row + 2 + df_no_valid_region.shape[0],
                output_3_start_row + len(output_3) + 2,
            ):
                for col_no in range(
                    title_3_start_col, title_3_start_col + output_3.shape[1]
                ):
                    cell = sheet1.cell(row=row_no, column=col_no)
                    cell.fill = yellow_fill

            title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
            title_4_start_col = 1
            output_4_start_row = title_4_start_row
            cell = sheet1.cell(
                row=title_4_start_row, column=title_4_start_col, value=title_4_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = blue_fill
            cell.font = Font(bold=True, size=12)
            sheet1.merge_cells(
                start_row=title_4_start_row,
                end_row=title_4_start_row,
                start_column=title_4_start_col,
                end_column=title_3_start_col + output_4.shape[1] - 1,
            )
            output_4.to_excel(
                writer,
                sheet_name="大区汇总信息",
                index=False,
                startrow=output_4_start_row,
            )

        # 打开已保存的 Excel 文件，进行进一步样式设置
        wb = load_workbook(excel_results_path)
        ws = wb["大区汇总信息"]

        # 设置列宽
        for i in range(9):
            ascii_value = ord("A")
            col_letter = chr(ascii_value + i)
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 30
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(excel_results_path)
    print()

    print()
    print("可疑经销商详细信息如下:")
    print("=" * 100)
    print()
    ids_suspicious = list(df_suspicious_dealers.BELONG_DEALER_NO)
    for id in ids_suspicious:
        df_dealer_dealer_results = df_dealer_results.loc[
            df_dealer_results.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
            df_total_scanning_locations.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_centroids = df_total_centroids.loc[
            df_total_centroids.dealer_id == id, :
        ].reset_index(drop=True)

        (
            title_1_str,
            title_2_str,
            output_2,
            title_3_str,
            output_3,
            title_4_str,
            output_4,
            title_5_str,
            output_5,
            title_6_str,
            output_6,
        ) = show_dealer_results_main(
            df_dealer_dealer_results,
            df_dealer_total_scanning_locations,
            df_dealer_total_centroids,
            dealer_scope_dict_path,
        )

        if save_results:
            excel_file_name_str = file_name_str
            map_file_name_str = f"{id}-{year_month_str}-品项：{product_group_id}"
            generate_single_dealer_sheet(
                results_files_folder_path,
                excel_file_name_str,
                map_file_name_str,
                id,
                title_1_str,
                title_2_str,
                output_2,
                title_3_str,
                output_3,
                title_4_str,
                output_4,
                title_5_str,
                output_5,
                title_6_str,
                output_6,
                mode="a",
                if_sheet_exists="overlay",
            )


def generate_all_dealers_results_main(
    df_dealer_results,
    df_total_scanning_locations,
    df_total_centroids,
    dealer_scope_dict_path,
    dealer_region_name,
    product_group_id,
    year_month_str,
    save_results=True,
):

    dealer_results_files_folder_path = (
        f"dealer_results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
    )
    os.makedirs(dealer_results_files_folder_path, exist_ok=True)
    if save_results:
        # Use glob to find all files in the directory
        files = glob.glob(
            os.path.join(dealer_results_files_folder_path, "*")
        )  # This matches all files and subdirectories
        for file in files:
            # Check if it's a file (not a subdirectory)
            if os.path.isfile(file):
                os.remove(file)

    df_dealer_results_within_archive = (
        df_dealer_results.loc[df_dealer_results["is_dealer_within_archive"] == 1,]
        .sort_values(by="BELONG_DEALER_NO")
        .reset_index(drop=True)
    )
    ids_dealers = list(df_dealer_results_within_archive.BELONG_DEALER_NO)

    for id in ids_dealers:
        df_dealer_dealer_results = df_dealer_results.loc[
            df_dealer_results.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
            df_total_scanning_locations.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_centroids = df_total_centroids.loc[
            df_total_centroids.dealer_id == id, :
        ].reset_index(drop=True)

        (
            title_1_str,
            title_2_str,
            output_2,
            title_3_str,
            output_3,
            title_4_str,
            output_4,
            title_5_str,
            output_5,
            title_6_str,
            output_6,
        ) = show_dealer_results_main(
            df_dealer_dealer_results,
            df_dealer_total_scanning_locations,
            df_dealer_total_centroids,
            dealer_scope_dict_path,
        )

        if save_results:
            excel_file_name_str = f"{id}-{year_month_str}-品项：{product_group_id}"
            map_file_name_str = f"{id}-{year_month_str}-品项：{product_group_id}"
            generate_single_dealer_sheet(
                dealer_results_files_folder_path,
                excel_file_name_str,
                map_file_name_str,
                id,
                title_1_str,
                title_2_str,
                output_2,
                title_3_str,
                output_3,
                title_4_str,
                output_4,
                title_5_str,
                output_5,
                title_6_str,
                output_6,
                mode="w",
            )


def generate_single_dealer_sheet_special(
    results_files_folder_path,
    excel_file_name_str,
    map_file_name_str,
    sheet_name,
    title_1_str,
    title_2_str,
    output_2,
    title_3_str,
    output_3,
    title_4_str,
    output_4,
    title_5_str,
    output_5,
    title_6_str,
    output_6,
    title_7_str,
    output_7,
    title_8_str,
    output_8,  # map1
    title_9_str,
    output_9,
    title_10_str,
    output_10,  # map2
    mode,
    **if_sheet_exists,
):
    excel_results_path = os.path.join(
        results_files_folder_path,
        f"{excel_file_name_str}.xlsx",
    )
    with pd.ExcelWriter(
        excel_results_path, engine="openpyxl", mode=mode, **if_sheet_exists
    ) as writer:
        # 创建or选择工作表
        if mode == "w":
            sheet1 = writer.book.create_sheet(sheet_name)
        if mode == "a":
            sheet1 = writer.book[sheet_name]
        blue_fill = PatternFill(
            start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
        )  # 浅蓝色
        yellow_fill = PatternFill(
            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
        )
        soft_red_fill = PatternFill(
            start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
        )

        # 写入标题 1
        title_1_start_row = 2
        title_1_start_col = 1
        cell = sheet1.cell(
            row=title_1_start_row, column=title_1_start_col, value=title_1_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=14)
        # 判断是否因 无有效经营范围 引发的可疑
        if output_3.loc[0, "无<有效>经营范围"] == "否":
            cell.fill = soft_red_fill
        else:
            cell.fill = yellow_fill
        sheet1.merge_cells(
            start_row=title_1_start_row,
            end_row=title_1_start_row,
            start_column=title_1_start_col,
            end_column=title_1_start_col + 10,
        )

        # 经营范围
        title_2_start_row = 4
        title_2_start_col = 1
        output_2_start_row = title_2_start_row
        cell = sheet1.cell(
            row=title_2_start_row, column=title_2_start_col, value=title_2_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_2_start_row,
            end_row=title_2_start_row,
            start_column=title_2_start_col,
            end_column=title_2_start_col + (output_2.shape[1]) - 1,
        )
        output_2.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_2_start_row
        )

        # 范围内经销商异地信息
        title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
        title_3_start_col = 1
        output_3_start_row = title_3_start_row
        cell = sheet1.cell(
            row=title_3_start_row, column=title_3_start_col, value=title_3_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_3_start_row,
            end_row=title_3_start_row,
            start_column=title_3_start_col,
            end_column=title_3_start_col + output_3.shape[1] - 1,
        )
        output_3.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_3_start_row
        )

        # 一级热点信息
        title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
        title_4_start_col = 1
        output_4_start_row = title_4_start_row
        cell = sheet1.cell(
            row=title_4_start_row, column=title_4_start_col, value=title_4_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_4_start_row,
            end_row=title_4_start_row,
            start_column=title_4_start_col,
            end_column=title_4_start_col + output_4.shape[1] - 1,
        )
        output_4.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_4_start_row
        )

        # 二级热点信息
        title_5_start_row = output_4_start_row + len(output_4) + 1 + 2
        title_5_start_col = 1
        output_5_start_row = title_5_start_row
        cell = sheet1.cell(
            row=title_5_start_row, column=title_5_start_col, value=title_5_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_5_start_row,
            end_row=title_5_start_row,
            start_column=title_5_start_col,
            end_column=title_5_start_col + output_5.shape[1] - 1,
        )
        output_5.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_5_start_row
        )

        # 可疑经销商开瓶城市统计表(大于五瓶的城市)
        title_6_start_row = output_5_start_row + len(output_5) + 1 + 2
        title_6_start_col = 1
        output_6_start_row = title_6_start_row
        cell = sheet1.cell(
            row=title_6_start_row, column=title_6_start_col, value=title_6_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, size=12)
        cell.fill = blue_fill
        sheet1.merge_cells(
            start_row=title_6_start_row,
            end_row=title_6_start_row,
            start_column=title_6_start_col,
            end_column=title_6_start_col + output_6.shape[1] - 1,
        )
        output_6.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_6_start_row
        )

        #  一级具体簇的信息
        title_7_start_row = output_6_start_row + len(output_6) + 1 + 2
        title_7_start_col = 1
        output_7_start_row = title_7_start_row
        cell = sheet1.cell(
            row=title_7_start_row, column=title_7_start_col, value=title_7_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = blue_fill
        cell.font = Font(bold=True, size=12)
        sheet1.merge_cells(
            start_row=title_7_start_row,
            end_row=title_7_start_row,
            start_column=title_7_start_col,
            end_column=title_7_start_col + output_7.shape[1] - 1,
        )
        # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
        output_7[["距本地热点总质心", "距本地点总质心"]] = output_7[
            ["距本地热点总质心", "距本地点总质心"]
        ].fillna("na")
        output_7.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_7_start_row
        )
        df_suspicious = output_7.loc[output_7["高度可疑"] == "是", :]
        for row_no in range(
            output_7_start_row + 2,
            output_7_start_row + df_suspicious.shape[0] + 2,
        ):
            for col_no in range(
                title_7_start_col, title_7_start_col + df_suspicious.shape[1]
            ):
                cell = sheet1.cell(row=row_no, column=col_no)
                cell.fill = soft_red_fill

        # 8 一级地图
        map_results_1_path = os.path.join(
            results_files_folder_path,
            f"{map_file_name_str}-SPARSE.html",
        )
        output_8.save(map_results_1_path)

        # 9 二级具体簇
        title_9_start_row = output_7_start_row + len(output_7) + 1 + 2
        title_9_start_col = 1
        output_9_start_row = title_9_start_row
        cell = sheet1.cell(
            row=title_9_start_row, column=title_9_start_col, value=title_9_str
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = blue_fill
        cell.font = Font(bold=True, size=12)
        sheet1.merge_cells(
            start_row=title_9_start_row,
            end_row=title_9_start_row,
            start_column=title_9_start_col,
            end_column=title_9_start_col + output_9.shape[1] - 1,
        )
        # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
        output_9[["距本地热点总质心", "距本地点总质心"]] = output_9[
            ["距本地热点总质心", "距本地点总质心"]
        ].fillna("na")
        output_9.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=output_9_start_row
        )
        df_suspicious = output_9.loc[output_9["高度可疑"] == "是", :]
        for row_no in range(
            output_9_start_row + 2,
            output_9_start_row + df_suspicious.shape[0] + 2,
        ):
            for col_no in range(
                title_9_start_col, title_9_start_col + df_suspicious.shape[1]
            ):
                cell = sheet1.cell(row=row_no, column=col_no)
                cell.fill = soft_red_fill

        # 10 map2 or str(没有地图)
        if type(output_10) != str:
            map_results_2_path = os.path.join(
                results_files_folder_path,
                f"{map_file_name_str}-DENSE.html",
            )
            output_10.save(map_results_2_path)

    # 打开已保存的 Excel 文件，进行进一步样式设置
    wb = load_workbook(excel_results_path)
    ws = wb[sheet_name]

    # 设置列宽
    for i in range(14):
        ascii_value = ord("A")
        col_letter = chr(ascii_value + i)
        ws.column_dimensions[col_letter].width = 20
    # ws.column_dimensions['A'].width = 30
    # ws.column_dimensions['B'].width = 40
    # ws.column_dimensions['C'].width = 30
    # 遍历每个单元格并设置居中对齐
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(excel_results_path)


def show_results_special_main(
    df_dealer_results,
    df_total_scanning_locations,
    df_total_centroids,
    df_suspicious_hotspots_parameters,
    df_dealer_results_dense,
    df_total_scanning_locations_dense,
    df_total_centroids_dense,
    df_suspicious_hotspots_parameters_dense,
    dealer_scope_dict_path,
    dealer_region_name,
    product_group_id,
    year_month_str,
    save_results=True,
):

    start_date_str, end_date_str = get_month_start_end(year_month_str)

    results_files_folder_path = (
        f"results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
    )
    os.makedirs(results_files_folder_path, exist_ok=True)

    rename_dict = {
        "radius": "簇半径",
        "min_samples": "簇内最少样本数",
        "dis_hotspots_c_t": "距本地热点总质心的距离阈值",
        "dis_points_c_t": "距本地扫码点总质心的距离阈值",
        "dis_border_t": "距边界最小距离阈值",
        "ratio_scanning_t": "热点扫码量占比阈值",
        "scanning_count_t": "热点扫码量阈值",
        "std_distance_t": "热点离散度阈值",
        "box_count_t": "紧密热点的箱数阈值",
        "BELONG_DEALER_NO": "经销商编码",
        "BELONG_DEALER_NAME": "经销商名称",
        "PRODUCT_GROUP_NAME": "品项名称",
    }

    df_model_parameters = df_suspicious_hotspots_parameters.copy().rename(
        columns=rename_dict
    )
    df_model_parameters_dense = df_suspicious_hotspots_parameters_dense.copy().rename(
        columns=rename_dict
    )
    df_model_parameters_dense.rename(
        columns={
            "簇半径": "二级分簇半径",
        }
    )

    df_suspicious_dealers_final = (
        df_dealer_results.loc[df_dealer_results["is_dealer_suspicious_final"] == 1, :]
        .sort_values(by="BELONG_DEALER_NO")
        .reset_index(drop=True)
    )

    product_group_name = df_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]
    main_title_str_1 = (
        "-" * 30
        + f"基于模型v1.0, ({start_date_str} - {end_date_str}),  ({dealer_region_name} - {product_group_name})的结果如下"
        + "-" * 30
    )
    print(main_title_str_1)
    print()

    main_title_str_2 = "--- 一级分簇模型参数 ---"
    print(main_title_str_2)
    print(
        tabulate(
            df_model_parameters, headers="keys", tablefmt="pretty", showindex=False
        )
    )
    print()

    main_title_str_3 = "--- 二级分簇模型参数 ---"
    print(main_title_str_3)
    print(
        tabulate(
            df_model_parameters_dense,
            headers="keys",
            tablefmt="pretty",
            showindex=False,
        )
    )
    print()

    (
        title_1_str,
        title_2_str,
        output_2,
        title_3_str,
        output_3,
        title_4_str,
        output_4,
    ) = show_region_short_results_special(
        df_dealer_results,
        df_total_scanning_locations,
        start_date_str,
        end_date_str,
        dealer_region_name,
    )

    if save_results:
        if os.path.exists(results_files_folder_path):
            # Use glob to find all files in the directory
            files = glob.glob(
                os.path.join(results_files_folder_path, "*")
            )  # This matches all files and subdirectories
            for file in files:
                # Check if it's a file (not a subdirectory)
                if os.path.isfile(file):
                    os.remove(file)

        file_name_str = f"{dealer_region_name}-{year_month_str}-{product_group_name}"
        excel_results_path = os.path.join(
            results_files_folder_path,
            f"{file_name_str}.xlsx",
        )

        # 输出大区汇总excel
        with pd.ExcelWriter(excel_results_path, engine="openpyxl") as writer:
            # 创建工作表

            sheet1 = writer.book.create_sheet("大区汇总信息")
            blue_fill = PatternFill(
                start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
            )  # 浅蓝色
            yellow_fill = PatternFill(
                start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
            )
            soft_red_fill = PatternFill(
                start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
            )
            hyperlink_font = Font(color="0000FF", underline="single")
            ids = list(output_3["经销商编码"])
            for i, id in enumerate(ids):
                writer.book.create_sheet(id)

            # 写入标题 1
            title_1_start_row = 1
            title_1_start_col = 1
            cell = sheet1.cell(
                row=title_1_start_row, column=title_1_start_col, value=title_1_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=14)
            sheet1.merge_cells(
                start_row=title_1_start_row,
                end_row=title_1_start_row,
                start_column=title_1_start_col,
                end_column=title_1_start_col + 11,
            )

            # 写入标题 2 和 output_2
            title_2_start_row = 3
            title_2_start_col = 1
            output_2_start_row = title_2_start_row
            cell = sheet1.cell(
                row=title_2_start_row, column=title_2_start_col, value=title_2_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=12)
            cell.fill = blue_fill
            sheet1.merge_cells(
                start_row=title_2_start_row,
                end_row=title_2_start_row,
                start_column=title_2_start_col,
                end_column=title_2_start_col + output_2.shape[1] - 1,
            )
            output_2.to_excel(
                writer,
                sheet_name="大区汇总信息",
                index=False,
                startrow=output_2_start_row,
            )

            title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
            title_3_start_col = 1
            output_3_start_row = title_3_start_row
            cell = sheet1.cell(
                row=title_3_start_row, column=title_3_start_col, value=title_3_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=12)
            cell.fill = blue_fill
            sheet1.merge_cells(
                start_row=title_3_start_row,
                end_row=title_3_start_row,
                start_column=title_3_start_col,
                end_column=title_3_start_col + output_3.shape[1] - 1,
            )
            output_3.to_excel(
                writer,
                sheet_name="大区汇总信息",
                index=False,
                startrow=output_3_start_row,
            )

            i = 0
            for row_no in range(
                output_3_start_row + 2, output_3_start_row + len(output_3) + 2
            ):
                cell_hyper = sheet1.cell(row=row_no, column=1)
                cell_hyper.hyperlink = f"#{ids[i]}!A1"
                cell_hyper.font = hyperlink_font
                i += 1
            # 并非 无任何有效经营范围而引发的可疑 填充红色
            df_no_valid_region = output_3.loc[output_3["无<有效>经营范围"] == "否", :]
            for row_no in range(
                output_3_start_row + 2,
                output_3_start_row + df_no_valid_region.shape[0] + 2,
            ):
                for col_no in range(
                    title_3_start_col, title_3_start_col + df_no_valid_region.shape[1]
                ):
                    cell = sheet1.cell(row=row_no, column=col_no)
                    cell.fill = soft_red_fill
            # 无任何有效经营范围引发的可疑 填充黄色
            for row_no in range(
                output_3_start_row + 2 + df_no_valid_region.shape[0],
                output_3_start_row + len(output_3) + 2,
            ):
                for col_no in range(
                    title_3_start_col, title_3_start_col + output_3.shape[1]
                ):
                    cell = sheet1.cell(row=row_no, column=col_no)
                    cell.fill = yellow_fill

            title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
            title_4_start_col = 1
            output_4_start_row = title_4_start_row
            cell = sheet1.cell(
                row=title_4_start_row, column=title_4_start_col, value=title_4_str
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = blue_fill
            cell.font = Font(bold=True, size=12)
            sheet1.merge_cells(
                start_row=title_4_start_row,
                end_row=title_4_start_row,
                start_column=title_4_start_col,
                end_column=title_3_start_col + output_4.shape[1] - 1,
            )
            output_4.to_excel(
                writer,
                sheet_name="大区汇总信息",
                index=False,
                startrow=output_4_start_row,
            )

            # 打开已保存的 Excel 文件，进行进一步样式设置
        wb = load_workbook(excel_results_path)
        ws = wb["大区汇总信息"]

        # 设置列宽
        for i in range(8):
            ascii_value = ord("A")
            col_letter = chr(ascii_value + i)
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 30
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(excel_results_path)
    print()

    print("可疑经销商详细信息如下:")
    print("=" * 100)
    print()
    ids_suspicious_total = list(df_suspicious_dealers_final.BELONG_DEALER_NO)
    for id in ids_suspicious_total:
        df_dealer_dealer_results = df_dealer_results.loc[
            df_dealer_results.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
            df_total_scanning_locations.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_centroids = df_total_centroids.loc[
            df_total_centroids.dealer_id == id, :
        ].reset_index(drop=True)

        df_dealer_dealer_results_dense = df_dealer_results_dense.loc[
            df_dealer_results_dense.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_scanning_locations_dense = (
            df_total_scanning_locations_dense.loc[
                df_total_scanning_locations_dense.BELONG_DEALER_NO == id, :
            ].reset_index(drop=True)
        )
        df_dealer_total_centroids_dense = df_total_centroids_dense.loc[
            df_total_centroids_dense.dealer_id == id, :
        ].reset_index(drop=True)

        (
            title_1_str,
            title_2_str,
            output_2,
            title_3_str,
            output_3,
            title_4_str,
            output_4,
            title_5_str,
            output_5,
            title_6_str,
            output_6,
            title_7_str,
            output_7,
            title_8_str,
            output_8,  # map1
            title_9_str,
            output_9,
            title_10_str,
            output_10,  # map2
        ) = show_dealer_results_special_main(
            df_dealer_dealer_results,
            df_dealer_total_scanning_locations,
            df_dealer_total_centroids,
            df_dealer_dealer_results_dense,
            df_dealer_total_scanning_locations_dense,
            df_dealer_total_centroids_dense,
            dealer_scope_dict_path,
        )

        if save_results:
            excel_file_name_str = file_name_str
            map_file_name_str = f"{id}-{year_month_str}-品项：{product_group_id}"
            generate_single_dealer_sheet_special(
                results_files_folder_path,
                excel_file_name_str,
                map_file_name_str,
                id,
                title_1_str,
                title_2_str,
                output_2,
                title_3_str,
                output_3,
                title_4_str,
                output_4,
                title_5_str,
                output_5,
                title_6_str,
                output_6,
                title_7_str,
                output_7,
                title_8_str,
                output_8,  # map1
                title_9_str,
                output_9,
                title_10_str,
                output_10,  # map2
                mode="a",
                if_sheet_exists="overlay",
            )


def generate_all_dealers_results_special_main(
    df_dealer_results,
    df_total_scanning_locations,
    df_total_centroids,
    df_dealer_results_dense,
    df_total_scanning_locations_dense,
    df_total_centroids_dense,
    dealer_scope_dict_path,
    dealer_region_name,
    product_group_id,
    year_month_str,
    save_results=True,
):

    dealer_results_files_folder_path = (
        f"dealer_results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
    )
    os.makedirs(dealer_results_files_folder_path, exist_ok=True)
    if save_results:
        # Use glob to find all files in the directory
        files = glob.glob(
            os.path.join(dealer_results_files_folder_path, "*")
        )  # This matches all files and subdirectories
        for file in files:
            # Check if it's a file (not a subdirectory)
            if os.path.isfile(file):
                os.remove(file)

    df_dealer_results_within_archive = (
        df_dealer_results.loc[df_dealer_results["is_dealer_within_archive"] == 1,]
        .sort_values(by="BELONG_DEALER_NO")
        .reset_index(drop=True)
    )
    ids_dealers = list(df_dealer_results_within_archive.BELONG_DEALER_NO)

    for id in ids_dealers:
        df_dealer_dealer_results = df_dealer_results.loc[
            df_dealer_results.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
            df_total_scanning_locations.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_centroids = df_total_centroids.loc[
            df_total_centroids.dealer_id == id, :
        ].reset_index(drop=True)

        df_dealer_dealer_results_dense = df_dealer_results_dense.loc[
            df_dealer_results_dense.BELONG_DEALER_NO == id, :
        ].reset_index(drop=True)
        df_dealer_total_scanning_locations_dense = (
            df_total_scanning_locations_dense.loc[
                df_total_scanning_locations_dense.BELONG_DEALER_NO == id, :
            ].reset_index(drop=True)
        )
        df_dealer_total_centroids_dense = df_total_centroids_dense.loc[
            df_total_centroids_dense.dealer_id == id, :
        ].reset_index(drop=True)

        (
            title_1_str,
            title_2_str,
            output_2,
            title_3_str,
            output_3,
            title_4_str,
            output_4,
            title_5_str,
            output_5,
            title_6_str,
            output_6,
            title_7_str,
            output_7,
            title_8_str,
            output_8,  # map1
            title_9_str,
            output_9,
            title_10_str,
            output_10,  # map2
        ) = show_dealer_results_special_main(
            df_dealer_dealer_results,
            df_dealer_total_scanning_locations,
            df_dealer_total_centroids,
            df_dealer_dealer_results_dense,
            df_dealer_total_scanning_locations_dense,
            df_dealer_total_centroids_dense,
            dealer_scope_dict_path,
        )

        if save_results:
            excel_file_name_str = f"{id}-{year_month_str}-品项：{product_group_id}"
            map_file_name_str = f"{id}-{year_month_str}-品项：{product_group_id}"
            generate_single_dealer_sheet_special(
                dealer_results_files_folder_path,
                excel_file_name_str,
                map_file_name_str,
                id,
                title_1_str,
                title_2_str,
                output_2,
                title_3_str,
                output_3,
                title_4_str,
                output_4,
                title_5_str,
                output_5,
                title_6_str,
                output_6,
                title_7_str,
                output_7,
                title_8_str,
                output_8,  # map1
                title_9_str,
                output_9,
                title_10_str,
                output_10,  # map2
                mode="w",
            )


# def show_results_main(
#     df_dealer_results,
#     df_total_scanning_locations,
#     df_total_centroids,
#     df_suspicious_hotspots_parameters,
#     dealer_scope_dict_path,
#     dealer_region_name,
#     product_group_id,
#     year_month_str,
#     save_results=False,
# ):
#     start_date_str, end_date_str = get_month_start_end(year_month_str)

#     results_files_folder_path = (
#         f"results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
#     )
#     # if os.path.exists(results_files_folder_path):
#     #     shutil.rmtree(results_files_folder_path)
#     os.makedirs(results_files_folder_path, exist_ok=True)

#     rename_dict = {
#         "radius": "簇半径",
#         "min_samples": "簇内最少样本数",
#         "dis_hotspots_c_t": "距本地热点总质心的距离阈值",
#         "dis_points_c_t": "距本地扫码点总质心的距离阈值",
#         "dis_border_t": "距边界最小距离阈值",
#         "ratio_scanning_t": "热点扫码量占比阈值",
#         "scanning_count_t": "热点扫码量阈值",
#         "std_distance_t": "热点离散度阈值",
#         "box_count_t": "紧密热点的箱数阈值",
#         "BELONG_DEALER_NO": "经销商编码",
#         "BELONG_DEALER_NAME": "经销商名称",
#         "PRODUCT_GROUP_NAME": "品项名称",
#     }

#     df_model_parameters = df_suspicious_hotspots_parameters.copy().rename(
#         columns=rename_dict
#     )
#     df_suspicious_dealers = df_dealer_results.loc[
#         df_dealer_results.is_dealer_suspicious == 1, :
#     ]
#     df_suspicious_dealers = df_suspicious_dealers.sort_values(
#         by="BELONG_DEALER_NO"
#     ).reset_index(drop=True)

#     product_group_name = df_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]
#     # product_group_name = df_suspicious_dealers.loc[0, "PRODUCT_GROUP_NAME"]

#     main_title_str_1 = (
#         "-" * 35
#         + f"基于模型v1.0, ({start_date_str} - {end_date_str}),  ({dealer_region_name} - {product_group_name})的结果如下"
#         + "-" * 35
#     )
#     print(main_title_str_1)
#     # display_forth_title(
#     #     main_title_str_1
#     # )
#     print()

#     # print("---模型参数---")
#     # print(tabulate(df_model_parameters, headers='keys', tablefmt='pretty', showindex=False))
#     main_title_str_2 = "--- 模型参数 ---"
#     print(main_title_str_2)
#     # display_fifth_title(main_title_str_2)
#     df_model_parameters_styled = (
#         df_model_parameters.style.format(
#             {
#                 "热点扫码量占比阈值": "{:.1f}",  # 浮动数值保留一位小数
#                 "热点离散度阈值": "{:.1f}",  # 浮动数值保留一位小数
#             }
#         )
#         .set_table_styles(
#             [
#                 {"selector": "th", "props": [("text-align", "center")]},  # 表头居中
#                 {"selector": "td", "props": [("text-align", "center")]},
#             ]  # 内容居中
#         )
#         .hide(axis="index")
#     )
#     print(
#         tabulate(
#             df_model_parameters, headers="keys", tablefmt="pretty", showindex=False
#         )
#     )
#     # display(df_model_parameters_styled)
#     print()

#     # show_region_short_results-----------------------
#     (
#         title_1_str,
#         title_2_str,
#         output_2,
#         title_3_str,
#         output_3,
#         title_4_str,
#         output_4,
#     ) = show_region_short_results(
#         df_dealer_results,
#         df_total_scanning_locations,
#         start_date_str,
#         end_date_str,
#         dealer_region_name,
#     )

#     # save excel part
#     if save_results:

#         if os.path.exists(results_files_folder_path):
#             # Use glob to find all files in the directory
#             files = glob.glob(os.path.join(results_files_folder_path, '*'))  # This matches all files and subdirectories

#             for file in files:
#                 # Check if it's a file (not a subdirectory)
#                 if os.path.isfile(file):
#                     os.remove(file)

#         excel_results_path = os.path.join(
#             results_files_folder_path,
#             f"{dealer_region_name}-{year_month_str}-{product_group_name}.xlsx",
#         )

#         # 输出普通大区汇总excel
#         with pd.ExcelWriter(excel_results_path, engine="openpyxl") as writer:
#             # 创建工作表

#             sheet1 = writer.book.create_sheet("大区汇总信息")
#             blue_fill = PatternFill(
#                 start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
#             )  # 浅蓝色
#             yellow_fill = PatternFill(
#                 start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
#             )
#             soft_red_fill = PatternFill(
#                 start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
#             )
#             hyperlink_font = Font(color="0000FF", underline="single")
#             ids = list(output_3["经销商编码"])
#             for i, id in enumerate(ids):
#                 writer.book.create_sheet(id)

#             # 写入标题 1
#             title_1_start_row = 1
#             title_1_start_col = 1
#             cell = sheet1.cell(
#                 row=title_1_start_row, column=title_1_start_col, value=title_1_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.font = Font(bold=True, size=14)
#             sheet1.merge_cells(
#                 start_row=title_1_start_row,
#                 end_row=title_1_start_row,
#                 start_column=title_1_start_col,
#                 end_column=title_1_start_col + 11,
#             )

#             # 写入标题 2 和 output_2
#             title_2_start_row = 3
#             title_2_start_col = 1
#             output_2_start_row = title_2_start_row
#             cell = sheet1.cell(
#                 row=title_2_start_row, column=title_2_start_col, value=title_2_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.font = Font(bold=True, size=12)
#             cell.fill = blue_fill
#             sheet1.merge_cells(
#                 start_row=title_2_start_row,
#                 end_row=title_2_start_row,
#                 start_column=title_2_start_col,
#                 end_column=title_2_start_col + output_2.shape[1] - 1,
#             )
#             output_2.to_excel(
#                 writer,
#                 sheet_name="大区汇总信息",
#                 index=False,
#                 startrow=output_2_start_row,
#             )

#             title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
#             title_3_start_col = 1
#             output_3_start_row = title_3_start_row
#             cell = sheet1.cell(
#                 row=title_3_start_row, column=title_3_start_col, value=title_3_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.font = Font(bold=True, size=12)
#             cell.fill = blue_fill
#             sheet1.merge_cells(
#                 start_row=title_3_start_row,
#                 end_row=title_3_start_row,
#                 start_column=title_3_start_col,
#                 end_column=title_3_start_col + output_3.shape[1] - 1,
#             )
#             output_3.to_excel(
#                 writer,
#                 sheet_name="大区汇总信息",
#                 index=False,
#                 startrow=output_3_start_row,
#             )

#             i = 0
#             for row_no in range(
#                 output_3_start_row + 2, output_3_start_row + len(output_3) + 2
#             ):
#                 cell_hyper = sheet1.cell(row=row_no, column=1)
#                 cell_hyper.hyperlink = f"#{ids[i]}!A1"
#                 cell_hyper.font = hyperlink_font
#                 i += 1
#             # 并非 无任何有效经营范围而引发的可疑 填充红色
#             df_no_valid_region = output_3.loc[output_3["无<有效>经营范围"] == "否", :]
#             for row_no in range(
#                 output_3_start_row + 2,
#                 output_3_start_row + df_no_valid_region.shape[0] + 2,
#             ):
#                 for col_no in range(
#                     title_3_start_col, title_3_start_col + df_no_valid_region.shape[1]
#                 ):
#                     cell = sheet1.cell(row=row_no, column=col_no)
#                     cell.fill = soft_red_fill
#             # 无任何有效经营范围引发的可疑 填充黄色
#             for row_no in range(
#                 output_3_start_row + 2 + df_no_valid_region.shape[0],
#                 output_3_start_row + len(output_3) + 2,
#             ):
#                 for col_no in range(
#                     title_3_start_col, title_3_start_col + output_3.shape[1]
#                 ):
#                     cell = sheet1.cell(row=row_no, column=col_no)
#                     cell.fill = yellow_fill


#             title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
#             title_4_start_col = 1
#             output_4_start_row = title_4_start_row
#             cell = sheet1.cell(
#                 row=title_4_start_row, column=title_4_start_col, value=title_4_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.fill = blue_fill
#             cell.font = Font(bold=True, size=12)
#             sheet1.merge_cells(
#                 start_row=title_4_start_row,
#                 end_row=title_4_start_row,
#                 start_column=title_4_start_col,
#                 end_column=title_3_start_col + output_4.shape[1] - 1,
#             )
#             output_4.to_excel(
#                 writer,
#                 sheet_name="大区汇总信息",
#                 index=False,
#                 startrow=output_4_start_row,
#             )

#         # 打开已保存的 Excel 文件，进行进一步样式设置
#         wb = load_workbook(excel_results_path)
#         ws = wb["大区汇总信息"]

#         # 设置列宽
#         for i in range(9):
#             ascii_value = ord("A")
#             col_letter = chr(ascii_value + i)
#             ws.column_dimensions[col_letter].width = 20
#         ws.column_dimensions["A"].width = 30
#         ws.column_dimensions["B"].width = 40
#         ws.column_dimensions["C"].width = 30
#         for row in ws.iter_rows():
#             for cell in row:
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#         wb.save(excel_results_path)
#     print()

#     print()
#     print("可疑经销商详细信息如下:")
#     print("=" * 100)
#     print()
#     ids_suspicious = list(df_suspicious_dealers.BELONG_DEALER_NO)
#     for id in ids_suspicious:
#         df_dealer_dealer_results = df_dealer_results.loc[
#             df_dealer_results.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
#             df_total_scanning_locations.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_centroids = df_total_centroids.loc[
#             df_total_centroids.dealer_id == id, :
#         ].reset_index(drop=True)

#         (
#             title_1_str,
#             title_2_str,
#             output_2,
#             title_3_str,
#             output_3,
#             title_4_str,
#             output_4,
#             title_5_str,
#             output_5,
#             title_6_str,
#             output_6,
#         ) = show_dealer_results_main(
#             df_dealer_dealer_results,
#             df_dealer_total_scanning_locations,
#             df_dealer_total_centroids,
#             dealer_scope_dict_path,
#         )

#         if save_results:
#             with pd.ExcelWriter(
#                 excel_results_path,
#                 engine="openpyxl",
#                 mode="a",
#                 if_sheet_exists="overlay",
#             ) as writer:
#                 # 创建工作表
#                 # sheet1 = writer.book.create_sheet(id)
#                 sheet_name = id
#                 sheet1 = writer.book[sheet_name]
#                 blue_fill = PatternFill(
#                     start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
#                 )  # 浅蓝色
#                 yellow_fill = PatternFill(
#                     start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
#                 )
#                 soft_red_fill = PatternFill(
#                     start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
#                 )
#                 # 写入标题 1
#                 title_1_start_row = 2
#                 title_1_start_col = 1
#                 cell = sheet1.cell(
#                     row=title_1_start_row, column=title_1_start_col, value=title_1_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=14)
#                 # 判断是否因 无有效经营范围 引发的可疑
#                 if output_3.loc[0, '无<有效>经营范围'] == '否':
#                     cell.fill = soft_red_fill
#                 else:
#                     cell.fill = yellow_fill
#                 sheet1.merge_cells(
#                     start_row=title_1_start_row,
#                     end_row=title_1_start_row,
#                     start_column=title_1_start_col,
#                     end_column=title_1_start_col + 10,
#                 )

#                 # 写入标题 2 和 output_2
#                 title_2_start_row = 4
#                 title_2_start_col = 1
#                 output_2_start_row = title_2_start_row
#                 cell = sheet1.cell(
#                     row=title_2_start_row, column=title_2_start_col, value=title_2_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_2_start_row,
#                     end_row=title_2_start_row,
#                     start_column=title_2_start_col,
#                     end_column=title_2_start_col + (output_2.shape[1]) - 1,
#                 )
#                 output_2.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_2_start_row
#                 )

#                 title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
#                 title_3_start_col = 1
#                 output_3_start_row = title_3_start_row
#                 cell = sheet1.cell(
#                     row=title_3_start_row, column=title_3_start_col, value=title_3_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_3_start_row,
#                     end_row=title_3_start_row,
#                     start_column=title_3_start_col,
#                     end_column=title_3_start_col + output_3.shape[1] - 1,
#                 )
#                 output_3.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_3_start_row
#                 )

#                 # 4 具体簇的信息
#                 title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
#                 title_4_start_col = 1
#                 output_4_start_row = title_4_start_row
#                 cell = sheet1.cell(
#                     row=title_4_start_row, column=title_4_start_col, value=title_4_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.fill = blue_fill
#                 cell.font = Font(bold=True, size=12)
#                 sheet1.merge_cells(
#                     start_row=title_4_start_row,
#                     end_row=title_4_start_row,
#                     start_column=title_4_start_col,
#                     end_column=title_4_start_col + output_4.shape[1] - 1,
#                 )
#                 # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
#                 output_4[['距本地热点总质心', '距本地点总质心']] = output_4[['距本地热点总质心', '距本地点总质心']].fillna('na')
#                 output_4.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_4_start_row
#                 )
#                 df_suspicious = output_4.loc[output_4["高度可疑"] == "是", :]
#                 for row_no in range(
#                     output_4_start_row + 2,
#                     output_4_start_row + df_suspicious.shape[0] + 2,
#                 ):
#                     for col_no in range(
#                         title_4_start_col, title_4_start_col + df_suspicious.shape[1]
#                     ):
#                         cell = sheet1.cell(row=row_no, column=col_no)
#                         cell.fill = soft_red_fill

#                 # 5
#                 title_5_start_row = output_4_start_row + len(output_4) + 1 + 2
#                 title_5_start_col = 1
#                 output_5_start_row = title_5_start_row
#                 cell = sheet1.cell(
#                     row=title_5_start_row, column=title_5_start_col, value=title_5_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_5_start_row,
#                     end_row=title_5_start_row,
#                     start_column=title_5_start_col,
#                     end_column=title_5_start_col + output_5.shape[1] - 1,
#                 )
#                 output_5.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_5_start_row
#                 )

#                 # 6 output是map
#                 map_results_path = os.path.join(
#                     results_files_folder_path,
#                     f"{id}-{year_month_str}-{product_group_name}.html",
#                 )
#                 output_6.save(map_results_path)

#             # 打开已保存的 Excel 文件，进行进一步样式设置
#             wb = load_workbook(excel_results_path)
#             ws = wb[id]

#             # 设置列宽
#             for i in range(14):
#                 ascii_value = ord("A")
#                 col_letter = chr(ascii_value + i)
#                 ws.column_dimensions[col_letter].width = 20
#             # ws.column_dimensions['A'].width = 30
#             # ws.column_dimensions['B'].width = 40
#             # ws.column_dimensions['C'].width = 30
#             # 遍历每个单元格并设置居中对齐
#             for row in ws.iter_rows():
#                 for cell in row:
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#             wb.save(excel_results_path)


# def show_results_special_main(
#     df_dealer_results,
#     df_total_scanning_locations,
#     df_total_centroids,
#     df_suspicious_hotspots_parameters,
#     df_dealer_results_dense,
#     df_total_scanning_locations_dense,
#     df_total_centroids_dense,
#     df_suspicious_hotspots_parameters_dense,
#     dealer_scope_dict_path,
#     dealer_region_name,
#     product_group_id,
#     year_month_str,
#     save_results=False,
# ):

#     start_date_str, end_date_str = get_month_start_end(year_month_str)

#     results_files_folder_path = (
#         f"results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
#     )
#     # if os.path.exists(results_files_folder_path):
#     #     shutil.rmtree(results_files_folder_path)
#     os.makedirs(results_files_folder_path, exist_ok=True)

#     rename_dict = {
#         "radius": "簇半径",
#         "min_samples": "簇内最少样本数",
#         "dis_hotspots_c_t": "距本地热点总质心的距离阈值",
#         "dis_points_c_t": "距本地扫码点总质心的距离阈值",
#         "dis_border_t": "距边界最小距离阈值",
#         "ratio_scanning_t": "热点扫码量占比阈值",
#         "scanning_count_t": "热点扫码量阈值",
#         "std_distance_t": "热点离散度阈值",
#         "box_count_t": "紧密热点的箱数阈值",
#         "BELONG_DEALER_NO": "经销商编码",
#         "BELONG_DEALER_NAME": "经销商名称",
#         "PRODUCT_GROUP_NAME": "品项名称",
#     }

#     df_model_parameters = df_suspicious_hotspots_parameters.copy().rename(
#         columns=rename_dict
#     )
#     df_model_parameters_dense = df_suspicious_hotspots_parameters_dense.copy().rename(
#         columns=rename_dict
#     )
#     df_model_parameters_dense.rename(
#         columns={
#             "簇半径": "二级分簇半径",
#         }
#     )

#     df_suspicious_dealers = (
#         df_dealer_results.loc[df_dealer_results["is_dealer_suspicious"] == 1, :]
#         .sort_values(by="BELONG_DEALER_NO")
#         .reset_index(drop=True)
#     )

#     df_suspicious_dealers_final = (
#         df_dealer_results.loc[df_dealer_results["is_dealer_suspicious_final"] == 1, :]
#         .sort_values(by="BELONG_DEALER_NO")
#         .reset_index(drop=True)
#     )
#     # product_group_name = df_suspicious_dealers.loc[0, "PRODUCT_GROUP_NAME"]
#     product_group_name = df_total_scanning_locations.loc[0, "PRODUCT_GROUP_NAME"]

#     main_title_str_1 = (
#         "-" * 30
#         + f"基于模型v1.0, ({start_date_str} - {end_date_str}),  ({dealer_region_name} - {product_group_name})的结果如下"
#         + "-" * 30
#     )
#     print(main_title_str_1)
#     # display_forth_title(
#     #     "-" * 30
#     #     + f"基于模型v1.0, ({start_date_str} - {end_date_str}),  ({dealer_region_name} - {product_group_name})的结果如下"
#     #     + "-" * 30
#     # )
#     print()

#     main_title_str_2 = "--- 一级分簇模型参数 ---"
#     print(main_title_str_2)
#     # display_fifth_title(main_title_str_2)
#     df_model_parameters_styled = (
#         df_model_parameters.style.format(
#             {
#                 "热点扫码量占比阈值": "{:.1f}",  # 浮动数值保留一位小数
#                 "热点离散度阈值": "{:.1f}",  # 浮动数值保留一位小数
#             }
#         )
#         .set_table_styles(
#             [
#                 {"selector": "th", "props": [("text-align", "center")]},  # 表头居中
#                 {"selector": "td", "props": [("text-align", "center")]},
#             ]  # 内容居中
#         )
#         .hide(axis="index")
#     )
#     print(
#         tabulate(
#             df_model_parameters, headers="keys", tablefmt="pretty", showindex=False
#         )
#     )
#     # display(df_model_parameters_styled)
#     print()

#     main_title_str_3 = "--- 二级分簇模型参数 ---"
#     print(main_title_str_3)
#     # display_fifth_title(main_title_str_3)
#     df_model_parameters_dense_styled = (
#         df_model_parameters_dense.style.format(
#             {
#                 "热点扫码量占比阈值": "{:.1f}",  # 浮动数值保留一位小数
#                 "热点离散度阈值": "{:.1f}",  # 浮动数值保留一位小数
#             }
#         )
#         .set_table_styles(
#             [
#                 {"selector": "th", "props": [("text-align", "center")]},  # 表头居中
#                 {"selector": "td", "props": [("text-align", "center")]},
#             ]  # 内容居中
#         )
#         .hide(axis="index")
#     )
#     print(
#         tabulate(
#             df_model_parameters_dense,
#             headers="keys",
#             tablefmt="pretty",
#             showindex=False,
#         )
#     )
#     # display(df_model_parameters_dense_styled)
#     print()

#     (
#         title_1_str,
#         title_2_str,
#         output_2,
#         title_3_str,
#         output_3,
#         title_4_str,
#         output_4,
#     ) = show_region_short_results_special(
#         df_dealer_results,
#         df_total_scanning_locations,
#         start_date_str,
#         end_date_str,
#         dealer_region_name,
#     )

#     if save_results:

#         if os.path.exists(results_files_folder_path):
#             # Use glob to find all files in the directory
#             files = glob.glob(os.path.join(results_files_folder_path, '*'))  # This matches all files and subdirectories
#             for file in files:
#                 # Check if it's a file (not a subdirectory)
#                 if os.path.isfile(file):
#                     os.remove(file)

#         excel_results_path = os.path.join(
#             results_files_folder_path,
#             f"{dealer_region_name}-{year_month_str}-{product_group_name}.xlsx",
#         )

#         # 输出普通大区汇总excel
#         with pd.ExcelWriter(excel_results_path, engine="openpyxl") as writer:
#             # 创建工作表

#             sheet1 = writer.book.create_sheet("大区汇总信息")
#             blue_fill = PatternFill(
#                 start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
#             )  # 浅蓝色
#             yellow_fill = PatternFill(
#                 start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
#             )
#             soft_red_fill = PatternFill(
#                 start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
#             )
#             hyperlink_font = Font(color="0000FF", underline="single")
#             ids = list(output_3["经销商编码"])
#             for i, id in enumerate(ids):
#                 writer.book.create_sheet(id)

#             # 写入标题 1
#             title_1_start_row = 1
#             title_1_start_col = 1
#             cell = sheet1.cell(
#                 row=title_1_start_row, column=title_1_start_col, value=title_1_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.font = Font(bold=True, size=14)
#             sheet1.merge_cells(
#                 start_row=title_1_start_row,
#                 end_row=title_1_start_row,
#                 start_column=title_1_start_col,
#                 end_column=title_1_start_col + 11,
#             )

#             # 写入标题 2 和 output_2
#             title_2_start_row = 3
#             title_2_start_col = 1
#             output_2_start_row = title_2_start_row
#             cell = sheet1.cell(
#                 row=title_2_start_row, column=title_2_start_col, value=title_2_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.font = Font(bold=True, size=12)
#             cell.fill = blue_fill
#             sheet1.merge_cells(
#                 start_row=title_2_start_row,
#                 end_row=title_2_start_row,
#                 start_column=title_2_start_col,
#                 end_column=title_2_start_col + output_2.shape[1] - 1,
#             )
#             output_2.to_excel(
#                 writer,
#                 sheet_name="大区汇总信息",
#                 index=False,
#                 startrow=output_2_start_row,
#             )

#             title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
#             title_3_start_col = 1
#             output_3_start_row = title_3_start_row
#             cell = sheet1.cell(
#                 row=title_3_start_row, column=title_3_start_col, value=title_3_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.font = Font(bold=True, size=12)
#             cell.fill = blue_fill
#             sheet1.merge_cells(
#                 start_row=title_3_start_row,
#                 end_row=title_3_start_row,
#                 start_column=title_3_start_col,
#                 end_column=title_3_start_col + output_3.shape[1] - 1,
#             )
#             output_3.to_excel(
#                 writer,
#                 sheet_name="大区汇总信息",
#                 index=False,
#                 startrow=output_3_start_row,
#             )

#             i = 0
#             for row_no in range(
#                 output_3_start_row + 2, output_3_start_row + len(output_3) + 2
#             ):
#                 cell_hyper = sheet1.cell(row=row_no, column=1)
#                 cell_hyper.hyperlink = f"#{ids[i]}!A1"
#                 cell_hyper.font = hyperlink_font
#                 i += 1
#             # 并非 无任何有效经营范围而引发的可疑 填充红色
#             df_no_valid_region = output_3.loc[output_3["无<有效>经营范围"] == "否", :]
#             for row_no in range(
#                 output_3_start_row + 2,
#                 output_3_start_row + df_no_valid_region.shape[0] + 2,
#             ):
#                 for col_no in range(
#                     title_3_start_col, title_3_start_col + df_no_valid_region.shape[1]
#                 ):
#                     cell = sheet1.cell(row=row_no, column=col_no)
#                     cell.fill = soft_red_fill
#             # 无任何有效经营范围引发的可疑 填充黄色
#             for row_no in range(
#                 output_3_start_row + 2 + df_no_valid_region.shape[0],
#                 output_3_start_row + len(output_3) + 2,
#             ):
#                 for col_no in range(
#                     title_3_start_col, title_3_start_col + output_3.shape[1]
#                 ):
#                     cell = sheet1.cell(row=row_no, column=col_no)
#                     cell.fill = yellow_fill

#             title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
#             title_4_start_col = 1
#             output_4_start_row = title_4_start_row
#             cell = sheet1.cell(
#                 row=title_4_start_row, column=title_4_start_col, value=title_4_str
#             )
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.fill = blue_fill
#             cell.font = Font(bold=True, size=12)
#             sheet1.merge_cells(
#                 start_row=title_4_start_row,
#                 end_row=title_4_start_row,
#                 start_column=title_4_start_col,
#                 end_column=title_3_start_col + output_4.shape[1] - 1,
#             )
#             output_4.to_excel(
#                 writer,
#                 sheet_name="大区汇总信息",
#                 index=False,
#                 startrow=output_4_start_row,
#             )

#             # 打开已保存的 Excel 文件，进行进一步样式设置
#         wb = load_workbook(excel_results_path)
#         ws = wb["大区汇总信息"]

#         # 设置列宽
#         for i in range(8):
#             ascii_value = ord("A")
#             col_letter = chr(ascii_value + i)
#             ws.column_dimensions[col_letter].width = 20
#         ws.column_dimensions["A"].width = 30
#         ws.column_dimensions["B"].width = 40
#         ws.column_dimensions["C"].width = 30
#         for row in ws.iter_rows():
#             for cell in row:
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#         wb.save(excel_results_path)
#     print()

#     print("可疑经销商详细信息如下:")
#     print("=" * 100)
#     print()
#     ids_suspicious_total = list(df_suspicious_dealers_final.BELONG_DEALER_NO)
#     for id in ids_suspicious_total:
#         df_dealer_dealer_results = df_dealer_results.loc[
#             df_dealer_results.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
#             df_total_scanning_locations.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_centroids = df_total_centroids.loc[
#             df_total_centroids.dealer_id == id, :
#         ].reset_index(drop=True)

#         df_dealer_dealer_results_dense = df_dealer_results_dense.loc[
#             df_dealer_results_dense.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_scanning_locations_dense = (
#             df_total_scanning_locations_dense.loc[
#                 df_total_scanning_locations_dense.BELONG_DEALER_NO == id, :
#             ].reset_index(drop=True)
#         )
#         df_dealer_total_centroids_dense = df_total_centroids_dense.loc[
#             df_total_centroids_dense.dealer_id == id, :
#         ].reset_index(drop=True)

#         (
#             title_1_str,
#             title_2_str,
#             output_2,
#             title_3_str,
#             output_3,
#             title_4_str,
#             output_4,
#             title_5_str,
#             output_5,
#             title_6_str,
#             output_6,
#             title_7_str,
#             output_7,
#             title_8_str,
#             output_8,  # map1
#             title_9_str,
#             output_9,
#             title_10_str,
#             output_10,  # map2
#         ) = show_dealer_results_special_main(
#             df_dealer_dealer_results,
#             df_dealer_total_scanning_locations,
#             df_dealer_total_centroids,
#             df_dealer_dealer_results_dense,
#             df_dealer_total_scanning_locations_dense,
#             df_dealer_total_centroids_dense,
#             dealer_scope_dict_path,
#         )

#         if save_results:
#             with pd.ExcelWriter(
#                 excel_results_path,
#                 engine="openpyxl",
#                 mode="a",
#                 if_sheet_exists="overlay",
#             ) as writer:
#                 # 创建工作表
#                 # sheet1 = writer.book.create_sheet(id)
#                 sheet_name = id
#                 sheet1 = writer.book[sheet_name]
#                 blue_fill = PatternFill(
#                     start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
#                 )  # 浅蓝色
#                 yellow_fill = PatternFill(
#                     start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
#                 )
#                 soft_red_fill = PatternFill(
#                     start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
#                 )

#                 # 写入标题 1
#                 title_1_start_row = 2
#                 title_1_start_col = 1
#                 cell = sheet1.cell(
#                     row=title_1_start_row, column=title_1_start_col, value=title_1_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=14)
#                 # 判断是否因 无有效经营范围 引发的可疑
#                 if output_3.loc[0, '无<有效>经营范围'] == '否':
#                     cell.fill = soft_red_fill
#                 else:
#                     cell.fill = yellow_fill
#                 sheet1.merge_cells(
#                     start_row=title_1_start_row,
#                     end_row=title_1_start_row,
#                     start_column=title_1_start_col,
#                     end_column=title_1_start_col + 10,
#                 )

#                 # 经营范围
#                 title_2_start_row = 4
#                 title_2_start_col = 1
#                 output_2_start_row = title_2_start_row
#                 cell = sheet1.cell(
#                     row=title_2_start_row, column=title_2_start_col, value=title_2_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_2_start_row,
#                     end_row=title_2_start_row,
#                     start_column=title_2_start_col,
#                     end_column=title_2_start_col + (output_2.shape[1]) - 1,
#                 )
#                 output_2.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_2_start_row
#                 )

#                 # 范围内经销商异地信息
#                 title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
#                 title_3_start_col = 1
#                 output_3_start_row = title_3_start_row
#                 cell = sheet1.cell(
#                     row=title_3_start_row, column=title_3_start_col, value=title_3_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_3_start_row,
#                     end_row=title_3_start_row,
#                     start_column=title_3_start_col,
#                     end_column=title_3_start_col + output_3.shape[1] - 1,
#                 )
#                 output_3.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_3_start_row
#                 )

#                 # 一级热点信息
#                 title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
#                 title_4_start_col = 1
#                 output_4_start_row = title_4_start_row
#                 cell = sheet1.cell(
#                     row=title_4_start_row, column=title_4_start_col, value=title_4_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_4_start_row,
#                     end_row=title_4_start_row,
#                     start_column=title_4_start_col,
#                     end_column=title_4_start_col + output_4.shape[1] - 1,
#                 )
#                 output_4.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_4_start_row
#                 )

#                 # 二级热点信息
#                 title_5_start_row = output_4_start_row + len(output_4) + 1 + 2
#                 title_5_start_col = 1
#                 output_5_start_row = title_5_start_row
#                 cell = sheet1.cell(
#                     row=title_5_start_row, column=title_5_start_col, value=title_5_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_5_start_row,
#                     end_row=title_5_start_row,
#                     start_column=title_5_start_col,
#                     end_column=title_5_start_col + output_5.shape[1] - 1,
#                 )
#                 output_5.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_5_start_row
#                 )

#                 # 可疑经销商开瓶城市统计表(大于五瓶的城市)
#                 title_6_start_row = output_5_start_row + len(output_5) + 1 + 2
#                 title_6_start_col = 1
#                 output_6_start_row = title_6_start_row
#                 cell = sheet1.cell(
#                     row=title_6_start_row, column=title_6_start_col, value=title_6_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_6_start_row,
#                     end_row=title_6_start_row,
#                     start_column=title_6_start_col,
#                     end_column=title_6_start_col + output_6.shape[1] - 1,
#                 )
#                 output_6.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_6_start_row
#                 )

#                 #  一级具体簇的信息
#                 title_7_start_row = output_6_start_row + len(output_6) + 1 + 2
#                 title_7_start_col = 1
#                 output_7_start_row = title_7_start_row
#                 cell = sheet1.cell(
#                     row=title_7_start_row, column=title_7_start_col, value=title_7_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.fill = blue_fill
#                 cell.font = Font(bold=True, size=12)
#                 sheet1.merge_cells(
#                     start_row=title_7_start_row,
#                     end_row=title_7_start_row,
#                     start_column=title_7_start_col,
#                     end_column=title_7_start_col + output_7.shape[1] - 1,
#                 )
#                 # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
#                 output_7[['距本地热点总质心', '距本地点总质心']] = output_7[['距本地热点总质心', '距本地点总质心']].fillna('na')
#                 output_7.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_7_start_row
#                 )
#                 df_suspicious = output_7.loc[output_7["高度可疑"] == "是", :]
#                 for row_no in range(
#                     output_7_start_row + 2,
#                     output_7_start_row + df_suspicious.shape[0] + 2,
#                 ):
#                     for col_no in range(
#                         title_7_start_col, title_7_start_col + df_suspicious.shape[1]
#                     ):
#                         cell = sheet1.cell(row=row_no, column=col_no)
#                         cell.fill = soft_red_fill

#                 # 8 一级地图
#                 map_results_1_path = os.path.join(
#                     results_files_folder_path,
#                     f"{id}-SPARSE-{year_month_str}-{product_group_name}.html",
#                 )
#                 output_8.save(map_results_1_path)

#                 # 9 二级具体簇
#                 title_9_start_row = output_7_start_row + len(output_7) + 1 + 2
#                 title_9_start_col = 1
#                 output_9_start_row = title_9_start_row
#                 cell = sheet1.cell(
#                     row=title_9_start_row, column=title_9_start_col, value=title_9_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.fill = blue_fill
#                 cell.font = Font(bold=True, size=12)
#                 sheet1.merge_cells(
#                     start_row=title_9_start_row,
#                     end_row=title_9_start_row,
#                     start_column=title_9_start_col,
#                     end_column=title_9_start_col + output_9.shape[1] - 1,
#                 )
#                 # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
#                 output_9[['距本地热点总质心', '距本地点总质心']] = output_9[['距本地热点总质心', '距本地点总质心']].fillna('na')
#                 output_9.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_9_start_row
#                 )
#                 df_suspicious = output_9.loc[output_9["高度可疑"] == "是", :]
#                 for row_no in range(
#                     output_9_start_row + 2,
#                     output_9_start_row + df_suspicious.shape[0] + 2,
#                 ):
#                     for col_no in range(
#                         title_9_start_col, title_9_start_col + df_suspicious.shape[1]
#                     ):
#                         cell = sheet1.cell(row=row_no, column=col_no)
#                         cell.fill = soft_red_fill

#                 # 10 map2 or str(没有地图)
#                 if type(output_10) != str:
#                     map_results_2_path = os.path.join(
#                         results_files_folder_path,
#                         f"{id}-DENSE-{year_month_str}-{product_group_name}.html",
#                     )
#                     output_10.save(map_results_2_path)

#             # 打开已保存的 Excel 文件，进行进一步样式设置
#             wb = load_workbook(excel_results_path)
#             ws = wb[id]

#             # 设置列宽
#             for i in range(14):
#                 ascii_value = ord("A")
#                 col_letter = chr(ascii_value + i)
#                 ws.column_dimensions[col_letter].width = 20
#             # ws.column_dimensions['A'].width = 30
#             # ws.column_dimensions['B'].width = 40
#             # ws.column_dimensions['C'].width = 30
#             # 遍历每个单元格并设置居中对齐
#             for row in ws.iter_rows():
#                 for cell in row:
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#             wb.save(excel_results_path)


# def generate_all_dealer_results(df_dealer_results, df_total_scanning_locations, df_total_centroids,
#                          dealer_scope_dict_path, dealer_region_name, product_group_id, year_month_str, save_results=True):
#     dealer_results_files_folder_path = f"dealer_results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
#     os.makedirs(dealer_results_files_folder_path, exist_ok=True)

#     df_dealer_results_within_archive = df_dealer_results.loc[df_dealer_results["is_dealer_within_archive"] == 1, ].sort_values(by="BELONG_DEALER_NO").reset_index(drop=True)
#     ids_dealers = list(df_dealer_results_within_archive.BELONG_DEALER_NO)

#     for id in ids_dealers:
#         df_dealer_dealer_results = df_dealer_results.loc[
#             df_dealer_results.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
#             df_total_scanning_locations.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_centroids = df_total_centroids.loc[
#             df_total_centroids.dealer_id == id, :
#         ].reset_index(drop=True)

#         (
#             title_1_str,
#             title_2_str,
#             output_2,
#             title_3_str,
#             output_3,
#             title_4_str,
#             output_4,
#             title_5_str,
#             output_5,
#             title_6_str,
#             output_6,
#         ) = show_dealer_results_main(
#             df_dealer_dealer_results,
#             df_dealer_total_scanning_locations,
#             df_dealer_total_centroids,
#             dealer_scope_dict_path,
#         )

#         if save_results:

#             excel_results_path = os.path.join(
#                 dealer_results_files_folder_path,
#                 f"{id}-{year_month_str}-品项：{product_group_id}.xlsx",
#             )
#             with pd.ExcelWriter(
#                 excel_results_path,
#                 engine="openpyxl",
#             ) as writer:
#                 # 创建工作表
#                 # sheet1 = writer.book.create_sheet(id)
#                 # sheet1 = writer.book[id]
#                 sheet_name = id
#                 sheet1 = writer.book.create_sheet(sheet_name)
#                 blue_fill = PatternFill(
#                     start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
#                 )  # 浅蓝色
#                 yellow_fill = PatternFill(
#                     start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
#                 )
#                 soft_red_fill = PatternFill(
#                     start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
#                 )
#                 # 写入标题 1
#                 title_1_start_row = 2
#                 title_1_start_col = 1
#                 cell = sheet1.cell(
#                     row=title_1_start_row, column=title_1_start_col, value=title_1_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=14)
#                 # 判断是否因 无有效经营范围 引发的可疑
#                 if output_3.loc[0, '无<有效>经营范围'] == '否':
#                     cell.fill = soft_red_fill
#                 else:
#                     cell.fill = yellow_fill
#                 sheet1.merge_cells(
#                     start_row=title_1_start_row,
#                     end_row=title_1_start_row,
#                     start_column=title_1_start_col,
#                     end_column=title_1_start_col + 10,
#                 )

#                 # 写入标题 2 和 output_2
#                 title_2_start_row = 4
#                 title_2_start_col = 1
#                 output_2_start_row = title_2_start_row
#                 cell = sheet1.cell(
#                     row=title_2_start_row, column=title_2_start_col, value=title_2_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_2_start_row,
#                     end_row=title_2_start_row,
#                     start_column=title_2_start_col,
#                     end_column=title_2_start_col + (output_2.shape[1]) - 1,
#                 )
#                 output_2.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_2_start_row
#                 )

#                 title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
#                 title_3_start_col = 1
#                 output_3_start_row = title_3_start_row
#                 cell = sheet1.cell(
#                     row=title_3_start_row, column=title_3_start_col, value=title_3_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_3_start_row,
#                     end_row=title_3_start_row,
#                     start_column=title_3_start_col,
#                     end_column=title_3_start_col + output_3.shape[1] - 1,
#                 )
#                 output_3.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_3_start_row
#                 )

#                 # 4 具体簇的信息
#                 title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
#                 title_4_start_col = 1
#                 output_4_start_row = title_4_start_row
#                 cell = sheet1.cell(
#                     row=title_4_start_row, column=title_4_start_col, value=title_4_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.fill = blue_fill
#                 cell.font = Font(bold=True, size=12)
#                 sheet1.merge_cells(
#                     start_row=title_4_start_row,
#                     end_row=title_4_start_row,
#                     start_column=title_4_start_col,
#                     end_column=title_4_start_col + output_4.shape[1] - 1,
#                 )
#                 # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
#                 output_4[['距本地热点总质心', '距本地点总质心']] = output_4[['距本地热点总质心', '距本地点总质心']].fillna('na')
#                 output_4.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_4_start_row
#                 )
#                 df_suspicious = output_4.loc[output_4["高度可疑"] == "是", :]
#                 for row_no in range(
#                     output_4_start_row + 2,
#                     output_4_start_row + df_suspicious.shape[0] + 2,
#                 ):
#                     for col_no in range(
#                         title_4_start_col, title_4_start_col + df_suspicious.shape[1]
#                     ):
#                         cell = sheet1.cell(row=row_no, column=col_no)
#                         cell.fill = soft_red_fill

#                 # 5
#                 title_5_start_row = output_4_start_row + len(output_4) + 1 + 2
#                 title_5_start_col = 1
#                 output_5_start_row = title_5_start_row
#                 cell = sheet1.cell(
#                     row=title_5_start_row, column=title_5_start_col, value=title_5_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_5_start_row,
#                     end_row=title_5_start_row,
#                     start_column=title_5_start_col,
#                     end_column=title_5_start_col + output_5.shape[1] - 1,
#                 )
#                 output_5.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_5_start_row
#                 )

#                 # 6 output是map
#                 map_results_path = os.path.join(
#                     dealer_results_files_folder_path,
#                     f"{id}-{year_month_str}-品项：{product_group_id}.html",
#                 )
#                 output_6.save(map_results_path)

#             # 打开已保存的 Excel 文件，进行进一步样式设置
#             wb = load_workbook(excel_results_path)
#             ws = wb[sheet_name]

#             # 设置列宽
#             for i in range(14):
#                 ascii_value = ord("A")
#                 col_letter = chr(ascii_value + i)
#                 ws.column_dimensions[col_letter].width = 20
#             # ws.column_dimensions['A'].width = 30
#             # ws.column_dimensions['B'].width = 40
#             # ws.column_dimensions['C'].width = 30
#             # 遍历每个单元格并设置居中对齐
#             for row in ws.iter_rows():
#                 for cell in row:
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#             wb.save(excel_results_path)

# def generate_all_dealer_results_special(
#     df_dealer_results,
#     df_total_scanning_locations,
#     df_total_centroids,
#     df_dealer_results_dense,
#     df_total_scanning_locations_dense,
#     df_total_centroids_dense,
#     dealer_scope_dict_path,
#     dealer_region_name,
#     product_group_id,
#     year_month_str,
#     save_results=True,
# ):
#     dealer_results_files_folder_path = f"dealer_results/{dealer_region_name}/{product_group_id}/{year_month_str}/"
#     os.makedirs(dealer_results_files_folder_path, exist_ok=True)

#     df_dealer_results_within_archive = df_dealer_results.loc[df_dealer_results["is_dealer_within_archive"] == 1, ].sort_values(by="BELONG_DEALER_NO").reset_index(drop=True)
#     ids_dealers = list(df_dealer_results_within_archive.BELONG_DEALER_NO)
#     for id in ids_dealers:
#         df_dealer_dealer_results = df_dealer_results.loc[
#             df_dealer_results.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_scanning_locations = df_total_scanning_locations.loc[
#             df_total_scanning_locations.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_centroids = df_total_centroids.loc[
#             df_total_centroids.dealer_id == id, :
#         ].reset_index(drop=True)

#         df_dealer_dealer_results_dense = df_dealer_results_dense.loc[
#             df_dealer_results_dense.BELONG_DEALER_NO == id, :
#         ].reset_index(drop=True)
#         df_dealer_total_scanning_locations_dense = (
#             df_total_scanning_locations_dense.loc[
#                 df_total_scanning_locations_dense.BELONG_DEALER_NO == id, :
#             ].reset_index(drop=True)
#         )
#         df_dealer_total_centroids_dense = df_total_centroids_dense.loc[
#             df_total_centroids_dense.dealer_id == id, :
#         ].reset_index(drop=True)

#         (
#             title_1_str,
#             title_2_str,
#             output_2,
#             title_3_str,
#             output_3,
#             title_4_str,
#             output_4,
#             title_5_str,
#             output_5,
#             title_6_str,
#             output_6,
#             title_7_str,
#             output_7,
#             title_8_str,
#             output_8,  # map1
#             title_9_str,
#             output_9,
#             title_10_str,
#             output_10,  # map2
#         ) = show_dealer_results_special_main(
#             df_dealer_dealer_results,
#             df_dealer_total_scanning_locations,
#             df_dealer_total_centroids,
#             df_dealer_dealer_results_dense,
#             df_dealer_total_scanning_locations_dense,
#             df_dealer_total_centroids_dense,
#             dealer_scope_dict_path,
#         )

#         if save_results:

#             excel_results_path = os.path.join(
#                 dealer_results_files_folder_path,
#                 f"{id}-{year_month_str}-品项：{product_group_id}.xlsx",
#             )
#             with pd.ExcelWriter(
#                 excel_results_path,
#                 engine="openpyxl",
#             ) as writer:
#                 # 创建工作表

#                 sheet_name = id
#                 sheet1 = writer.book.create_sheet(sheet_name)
#                 # sheet1 = writer.book[sheet_name]
#                 blue_fill = PatternFill(
#                     start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
#                 )  # 浅蓝色
#                 yellow_fill = PatternFill(
#                     start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
#                 )
#                 soft_red_fill = PatternFill(
#                     start_color="FAD1D1", end_color="FAD1D1", fill_type="solid"
#                 )

#                 # 写入标题 1
#                 title_1_start_row = 2
#                 title_1_start_col = 1
#                 cell = sheet1.cell(
#                     row=title_1_start_row, column=title_1_start_col, value=title_1_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=14)
#                 # 判断是否因 无有效经营范围 引发的可疑
#                 if output_3.loc[0, '无<有效>经营范围'] == '否':
#                     cell.fill = soft_red_fill
#                 else:
#                     cell.fill = yellow_fill
#                 sheet1.merge_cells(
#                     start_row=title_1_start_row,
#                     end_row=title_1_start_row,
#                     start_column=title_1_start_col,
#                     end_column=title_1_start_col + 10,
#                 )

#                 # 经营范围
#                 title_2_start_row = 4
#                 title_2_start_col = 1
#                 output_2_start_row = title_2_start_row
#                 cell = sheet1.cell(
#                     row=title_2_start_row, column=title_2_start_col, value=title_2_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_2_start_row,
#                     end_row=title_2_start_row,
#                     start_column=title_2_start_col,
#                     end_column=title_2_start_col + (output_2.shape[1]) - 1,
#                 )
#                 output_2.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_2_start_row
#                 )

#                 # 范围内经销商异地信息
#                 title_3_start_row = output_2_start_row + len(output_2) + 1 + 2
#                 title_3_start_col = 1
#                 output_3_start_row = title_3_start_row
#                 cell = sheet1.cell(
#                     row=title_3_start_row, column=title_3_start_col, value=title_3_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_3_start_row,
#                     end_row=title_3_start_row,
#                     start_column=title_3_start_col,
#                     end_column=title_3_start_col + output_3.shape[1] - 1,
#                 )
#                 output_3.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_3_start_row
#                 )

#                 # 一级热点信息
#                 title_4_start_row = output_3_start_row + len(output_3) + 1 + 2
#                 title_4_start_col = 1
#                 output_4_start_row = title_4_start_row
#                 cell = sheet1.cell(
#                     row=title_4_start_row, column=title_4_start_col, value=title_4_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_4_start_row,
#                     end_row=title_4_start_row,
#                     start_column=title_4_start_col,
#                     end_column=title_4_start_col + output_4.shape[1] - 1,
#                 )
#                 output_4.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_4_start_row
#                 )

#                 # 二级热点信息
#                 title_5_start_row = output_4_start_row + len(output_4) + 1 + 2
#                 title_5_start_col = 1
#                 output_5_start_row = title_5_start_row
#                 cell = sheet1.cell(
#                     row=title_5_start_row, column=title_5_start_col, value=title_5_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_5_start_row,
#                     end_row=title_5_start_row,
#                     start_column=title_5_start_col,
#                     end_column=title_5_start_col + output_5.shape[1] - 1,
#                 )
#                 output_5.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_5_start_row
#                 )

#                 # 可疑经销商开瓶城市统计表(大于五瓶的城市)
#                 title_6_start_row = output_5_start_row + len(output_5) + 1 + 2
#                 title_6_start_col = 1
#                 output_6_start_row = title_6_start_row
#                 cell = sheet1.cell(
#                     row=title_6_start_row, column=title_6_start_col, value=title_6_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.font = Font(bold=True, size=12)
#                 cell.fill = blue_fill
#                 sheet1.merge_cells(
#                     start_row=title_6_start_row,
#                     end_row=title_6_start_row,
#                     start_column=title_6_start_col,
#                     end_column=title_6_start_col + output_6.shape[1] - 1,
#                 )
#                 output_6.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_6_start_row
#                 )

#                 #  一级具体簇的信息
#                 title_7_start_row = output_6_start_row + len(output_6) + 1 + 2
#                 title_7_start_col = 1
#                 output_7_start_row = title_7_start_row
#                 cell = sheet1.cell(
#                     row=title_7_start_row, column=title_7_start_col, value=title_7_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.fill = blue_fill
#                 cell.font = Font(bold=True, size=12)
#                 sheet1.merge_cells(
#                     start_row=title_7_start_row,
#                     end_row=title_7_start_row,
#                     start_column=title_7_start_col,
#                     end_column=title_7_start_col + output_7.shape[1] - 1,
#                 )
#                 # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
#                 output_7[['距本地热点总质心', '距本地点总质心']] = output_7[['距本地热点总质心', '距本地点总质心']].fillna('na')
#                 output_7.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_7_start_row
#                 )
#                 df_suspicious = output_7.loc[output_7["高度可疑"] == "是", :]
#                 for row_no in range(
#                     output_7_start_row + 2,
#                     output_7_start_row + df_suspicious.shape[0] + 2,
#                 ):
#                     for col_no in range(
#                         title_7_start_col, title_7_start_col + df_suspicious.shape[1]
#                     ):
#                         cell = sheet1.cell(row=row_no, column=col_no)
#                         cell.fill = soft_red_fill

#                 # 8 一级地图
#                 map_results_1_path = os.path.join(
#                     dealer_results_files_folder_path,
#                     f"{id}-SPARSE-{year_month_str}-品项：{product_group_id}.html",
#                 )
#                 output_8.save(map_results_1_path)

#                 # 9 二级具体簇
#                 title_9_start_row = output_7_start_row + len(output_7) + 1 + 2
#                 title_9_start_col = 1
#                 output_9_start_row = title_9_start_row
#                 cell = sheet1.cell(
#                     row=title_9_start_row, column=title_9_start_col, value=title_9_str
#                 )
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.fill = blue_fill
#                 cell.font = Font(bold=True, size=12)
#                 sheet1.merge_cells(
#                     start_row=title_9_start_row,
#                     end_row=title_9_start_row,
#                     start_column=title_9_start_col,
#                     end_column=title_9_start_col + output_9.shape[1] - 1,
#                 )
#                 # 具体簇的信息中 距离质心的距离如果为 np.nan 在输出excel时替换成'na'(显示提示'不适用')
#                 output_9[['距本地热点总质心', '距本地点总质心']] = output_9[['距本地热点总质心', '距本地点总质心']].fillna('na')
#                 output_9.to_excel(
#                     writer, sheet_name=sheet_name, index=False, startrow=output_9_start_row
#                 )
#                 df_suspicious = output_9.loc[output_9["高度可疑"] == "是", :]
#                 for row_no in range(
#                     output_9_start_row + 2,
#                     output_9_start_row + df_suspicious.shape[0] + 2,
#                 ):
#                     for col_no in range(
#                         title_9_start_col, title_9_start_col + df_suspicious.shape[1]
#                     ):
#                         cell = sheet1.cell(row=row_no, column=col_no)
#                         cell.fill = soft_red_fill

#                 # 10 map2 or str(没有地图)
#                 if type(output_10) != str:
#                     map_results_2_path = os.path.join(
#                         dealer_results_files_folder_path,
#                         f"{id}-DENSE-{year_month_str}-品项：{product_group_id}.html",
#                     )
#                     output_10.save(map_results_2_path)

#             # 打开已保存的 Excel 文件，进行进一步样式设置
#             wb = load_workbook(excel_results_path)
#             ws = wb[sheet_name]

#             # 设置列宽
#             for i in range(14):
#                 ascii_value = ord("A")
#                 col_letter = chr(ascii_value + i)
#                 ws.column_dimensions[col_letter].width = 20
#             # ws.column_dimensions['A'].width = 30
#             # ws.column_dimensions['B'].width = 40
#             # ws.column_dimensions['C'].width = 30
#             # 遍历每个单元格并设置居中对齐
#             for row in ws.iter_rows():
#                 for cell in row:
#                     cell.alignment = Alignment(horizontal="center", vertical="center")
#             wb.save(excel_results_path)
